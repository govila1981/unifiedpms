"""
Enhanced Deliverables Calculator Module
Integrates Excel formatting and Bloomberg formulas from the working version
Calculates physical delivery obligations with sophisticated Excel output
Uses centralized PriceManager for all price data
"""

import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

# Import centralized price manager
try:
    from price_manager import PriceManager
except ImportError:
    logger.warning("PriceManager not found - prices will need to be provided")
    PriceManager = None

logger = logging.getLogger(__name__)


class DeliverableCalculator:
    """Enhanced calculator for physical deliverables with better Excel output"""

    def __init__(self, usdinr_rate: float = 88.0):
        self.usdinr_rate = usdinr_rate
        self.wb = None

        # Initialize PriceManager if available
        if PriceManager:
            self.price_manager = PriceManager()
        else:
            self.price_manager = None
        
        # Excel styles
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.group_font = Font(bold=True, size=10)
        self.group_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        self.grand_total_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        # Number formats
        self.price_format = '#,##0.00'
        self.deliv_format = '#,##0'
        self.iv_format = '#,##0'
        self.percent_format = '0.00%'
    
    def calculate_deliverables_from_dataframe(self, positions_df: pd.DataFrame, 
                                             prices: Dict[str, float] = None) -> pd.DataFrame:
        """
        Calculate deliverables from a positions DataFrame
        Returns DataFrame for compatibility with existing code
        """
        if positions_df.empty:
            return pd.DataFrame()
        
        deliverables_data = []
        
        for idx, row in positions_df.iterrows():
            ticker = row.get('Ticker', '')
            symbol = row.get('Symbol', '')
            security_type = row.get('Security_Type', '')
            strike = float(row.get('Strike', 0))
            lots = float(row.get('Lots', 0))
            lot_size = int(row.get('Lot_Size', 1))
            
            # Get underlying ticker (for price lookup)
            underlying = row.get('Underlying', '')

            # Get price if available
            spot_price = 0

            # For options and futures, we need the underlying asset price
            if security_type in ['Call', 'Put', 'Futures']:
                if prices:
                    # For indices, symbol might be more accurate than underlying
                    # Check if this is an index option (NMIDSELP, NIFTY, etc)
                    is_index = (symbol and ('NMIDSELP' in symbol.upper() or 'NIFTY' in symbol.upper())) or \
                              ('INDEX' in str(row.get('Ticker', '')).upper()) or \
                              ('NMIDSELP' in str(row.get('Ticker', '')).upper())

                    if is_index:
                        # Try symbol first for index options
                        spot_price = prices.get(symbol, 0)
                        # Also try without spaces
                        if spot_price == 0 and symbol:
                            symbol_clean = symbol.replace(' ', '')
                            spot_price = prices.get(symbol_clean, 0)

                    # If not found or not index, try underlying
                    if spot_price == 0:
                        spot_price = prices.get(underlying, 0)

                    # If underlying price not found by name, try other fields
                    if spot_price == 0 and underlying:
                        # Try variations of the underlying symbol
                        for key in prices:
                            if underlying.upper() in key.upper() or key.upper() in underlying.upper():
                                spot_price = prices[key]
                                break

                    # Last resort - try symbol/ticker if they might be the underlying
                    if spot_price == 0:
                        spot_price = prices.get(symbol, prices.get(ticker, 0))

                elif self.price_manager:
                    # Use centralized price manager if no prices provided
                    spot_price = self.price_manager.get_price(underlying)
                    if not spot_price and underlying:
                        # Try variations
                        for suffix in ['', '.NS', '.BO', ' IS Equity', ' INDEX']:
                            test_symbol = underlying + suffix
                            spot_price = self.price_manager.get_price(test_symbol)
                            if spot_price:
                                break

                    if not spot_price:
                        spot_price = self.price_manager.get_price(symbol)
                    if not spot_price:
                        spot_price = self.price_manager.get_price(ticker)
                    if not spot_price:
                        spot_price = 0
            else:
                # For other security types, use direct price
                if prices:
                    spot_price = prices.get(symbol, prices.get(ticker, 0))
                elif self.price_manager:
                    spot_price = self.price_manager.get_price(symbol)
                    if not spot_price:
                        spot_price = self.price_manager.get_price(ticker)
                    if not spot_price:
                        spot_price = 0
            
            # Calculate deliverable
            if security_type == 'Futures':
                # Futures always deliver regardless of price
                deliverable = lots
            elif security_type == 'Call':
                if spot_price > 0 and spot_price > strike:
                    # ITM Call - will be exercised
                    deliverable = lots
                elif spot_price > 0:
                    # OTM Call - expires worthless
                    deliverable = 0
                else:
                    # No price available - assume 0 (cannot determine ITM/OTM)
                    deliverable = 0
                    logger.warning(f"No price for {symbol} ({ticker}), assuming 0 deliverable for Call")
            elif security_type == 'Put':
                if spot_price > 0 and spot_price < strike:
                    # ITM Put - will be exercised
                    deliverable = -lots
                elif spot_price > 0:
                    # OTM Put - expires worthless
                    deliverable = 0
                else:
                    # No price available - assume 0 (cannot determine ITM/OTM)
                    deliverable = 0
                    logger.warning(f"No price for {symbol} ({ticker}), assuming 0 deliverable for Put")
            else:
                deliverable = 0
            
            # Calculate intrinsic value
            iv = self._calculate_intrinsic_value(security_type, spot_price, strike, lots, lot_size)
            
            deliverables_data.append({
                'Ticker': ticker,
                'Symbol': symbol,
                'Underlying': underlying,
                'Security_Type': security_type,
                'Strike': strike,
                'Lots': lots,
                'Lot_Size': lot_size,
                'Spot_Price': spot_price if spot_price > 0 else 'N/A',
                'Deliverable_Lots': deliverable if deliverable is not None else 'Price Required',
                'Deliverable_Qty': (deliverable * lot_size) if deliverable is not None else 'Price Required',
                'Intrinsic_Value_INR': iv if spot_price > 0 else 'N/A',
                'Intrinsic_Value_USD': (iv / self.usdinr_rate) if spot_price > 0 else 'N/A'
            })
        
        return pd.DataFrame(deliverables_data)
    
    def _calculate_intrinsic_value(self, security_type: str, spot_price: float, 
                                  strike: float, lots: float, lot_size: int) -> float:
        """Calculate intrinsic value for an option position"""
        if security_type == 'Futures' or spot_price <= 0:
            return 0
        
        if security_type == 'Call':
            if spot_price > strike:
                return lots * lot_size * (spot_price - strike)
        elif security_type == 'Put':
            if spot_price < strike:
                return lots * lot_size * (strike - spot_price)
        
        return 0
    
    def generate_deliverables_report(self, 
                                    starting_positions_df: pd.DataFrame,
                                    final_positions_df: pd.DataFrame,
                                    prices: Dict[str, float],
                                    output_file: str,
                                    report_type: str = "TRADE_PROCESSING") -> str:
        """
        Generate comprehensive deliverables report with Bloomberg formulas
        Following the exact format from the working delivery calculator
        """
        self.wb = Workbook()
        self.prices = prices  # Store prices as instance variable for access by other methods
        
        # Remove default sheet
        if self.wb.active:
            self.wb.remove(self.wb.active)
        
        # Process positions into Position-like objects for compatibility
        pre_positions = self._convert_to_positions(starting_positions_df)
        post_positions = self._convert_to_positions(final_positions_df)
        
        # Combine all positions for master sheets
        all_positions = self._combine_positions(pre_positions, post_positions)
        
        # Write sheets for PRE and POST separately
        if pre_positions:
            self._write_master_sheet(pre_positions, prices, "PRE", "pre")
            self._write_iv_master_sheet(pre_positions, prices, "PRE", "pre")

            expiries = list(set(p['expiry'] for p in pre_positions))
            for expiry in sorted(expiries):
                self._write_expiry_sheet(expiry, pre_positions, prices, "PRE", "pre")

        if post_positions:
            self._write_master_sheet(post_positions, prices, "POST", "post")
            self._write_iv_master_sheet(post_positions, prices, "POST", "post")

            expiries = list(set(p['expiry'] for p in post_positions))
            for expiry in sorted(expiries):
                self._write_expiry_sheet(expiry, post_positions, prices, "POST", "post")
            
            # Write comparison sheet
            self._write_comparison_sheet(pre_positions, post_positions, prices)
        
        # Write all positions sheet
        self._write_all_positions_sheet(pre_positions, "Pre_Trade_Positions")
        self._write_all_positions_sheet(post_positions, "Post_Trade_Positions")
        
        # Save workbook
        self.wb.save(output_file)
        logger.info(f"Deliverables report saved: {output_file}")
        return output_file
    
    def _convert_to_positions(self, positions_df: pd.DataFrame) -> List[Dict]:
        """Convert DataFrame to position dictionaries for processing"""
        positions = []
        
        if positions_df.empty:
            return positions
        
        for idx, row in positions_df.iterrows():
            # Extract underlying from Ticker if Underlying column doesn't exist
            ticker = row.get('Ticker', '')
            underlying = row.get('Underlying', '')
            
            if not underlying:
                # Try to extract from Bloomberg ticker
                if 'IS Equity' in ticker:
                    underlying = ticker.split('=')[0] + ' IS Equity'
                elif 'Index' in ticker:
                    underlying = ticker.split(' ')[0] + ' Index'
                else:
                    underlying = ticker
            
            positions.append({
                'underlying': underlying,
                'ticker': ticker,
                'symbol': row.get('Symbol', ''),
                'expiry': pd.to_datetime(row.get('Expiry', datetime.now())),
                'lots': float(row.get('Lots', 0)),
                'security_type': row.get('Security_Type', ''),
                'strike': float(row.get('Strike', 0)),
                'lot_size': int(row.get('Lot_Size', 1))
            })
        
        return positions
    
    def _combine_positions(self, pre_positions: List[Dict], post_positions: List[Dict]) -> List[Dict]:
        """Combine pre and post positions, keeping unique tickers"""
        combined = {}
        
        for pos in pre_positions + post_positions:
            ticker = pos['ticker']
            if ticker not in combined:
                combined[ticker] = pos
        
        return list(combined.values())
    
    def _write_master_sheet(self, positions: List[Dict], prices: Dict[str, float], prefix: str, position_type: str = ""):
        """Write master sheet with all expiries INCLUDING BLOOMBERG FORMULAS"""
        sheet_name = f"{prefix}_Master_All_Expiries"
        ws = self.wb.create_sheet(sheet_name)

        # Store position type to filter correctly
        self.current_position_type = position_type
        
        # Headers matching the working version exactly
        headers = [
            "Underlying", "Symbol", "Expiry", "Position", "Type", "Strike",
            "System Deliverable", "Override Deliverable", "System Price",
            "Override Price", "BBG Price", "BBG Deliverable",
            "", "System Price -%", "System Deliv -%", "System Price +%", "System Deliv +%",
            "BBG Price -%", "BBG Deliv -%", "BBG Price +%", "BBG Deliv +%"
        ]
        
        # Add sensitivity percentage input in M1
        ws.cell(row=1, column=13, value=10)  # Default 10% sensitivity
        ws.cell(row=1, column=13).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            if col != 13:  # Skip the separator column
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Group positions by underlying
        grouped = self._group_by_underlying(positions)
        
        current_row = 3
        
        for underlying in sorted(grouped.keys()):
            underlying_positions = grouped[underlying]
            group_start_row = current_row
            
            # Write group header
            cell = ws.cell(row=current_row, column=1, value=underlying)
            cell.font = self.group_font
            
            # Apply group formatting
            for col in range(1, 22):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = self.group_fill
                cell.border = self.border
            
            # Get spot price from Yahoo (try different keys)
            spot_price = None
            if prices:
                # Try underlying first
                spot_price = prices.get(underlying)
                # If not found, try the symbol from first position
                if not spot_price and underlying_positions:
                    first_symbol = underlying_positions[0].get('symbol', '')
                    spot_price = prices.get(first_symbol)
                # If still not found, check if underlying has a space and try first part
                if not spot_price and ' ' in underlying:
                    base_ticker = underlying.split(' ')[0]
                    spot_price = prices.get(base_ticker)
            
            # System price in group header (Yahoo price)
            if spot_price:
                ws.cell(row=current_row, column=9, value=spot_price).number_format = self.price_format
            else:
                ws.cell(row=current_row, column=9, value="").number_format = self.price_format
            
            # Override price (blank for user input)
            ws.cell(row=current_row, column=10, value="").number_format = self.price_format
            
            # Bloomberg price formula
            bbg_formula = f'=BDP(A{current_row},"PX_LAST")'
            ws.cell(row=current_row, column=11, value=bbg_formula).number_format = self.price_format
            
            # Sensitivity price formulas for group header
            ws.cell(row=current_row, column=14, 
                value=f"=IF($M$1<>\"\",I{current_row}*(1-$M$1/100),\"\")").number_format = self.price_format
            ws.cell(row=current_row, column=16,
                value=f"=IF($M$1<>\"\",I{current_row}*(1+$M$1/100),\"\")").number_format = self.price_format
            ws.cell(row=current_row, column=18,
                value=f"=IF($M$1<>\"\",K{current_row}*(1-$M$1/100),\"\")").number_format = self.price_format
            ws.cell(row=current_row, column=20,
                value=f"=IF($M$1<>\"\",K{current_row}*(1+$M$1/100),\"\")").number_format = self.price_format
            
            detail_rows = []
            current_row += 1
            
            # Write detail rows
            for pos in sorted(underlying_positions, key=lambda x: (x['expiry'], x['strike'])):
                ws.cell(row=current_row, column=2, value=pos['ticker'])
                ws.cell(row=current_row, column=3, value=pos['expiry'].strftime('%d/%m/%Y'))
                ws.cell(row=current_row, column=4, value=pos['lots'])
                ws.cell(row=current_row, column=5, value=pos['security_type'])
                
                if pos['strike'] > 0:
                    ws.cell(row=current_row, column=6, value=pos['strike']).number_format = self.price_format
                
                # Deliverable formulas
                for col_idx, price_col in [(7, "I"), (8, "J"), (12, "K")]:
                    formula = self._create_deliverable_formula(
                        current_row, group_start_row, pos['security_type'], 
                        pos['strike'], pos['lots'], price_col
                    )
                    ws.cell(row=current_row, column=col_idx, value=formula).number_format = self.deliv_format
                
                # Sensitivity deliverable formulas
                for col_idx, price_col in [(15, "N"), (17, "P"), (19, "R"), (21, "T")]:
                    base_formula = self._create_deliverable_formula(
                        current_row, group_start_row, pos['security_type'],
                        pos['strike'], pos['lots'], price_col
                    )
                    cell = ws.cell(row=current_row, column=col_idx, 
                        value=f"=IF($M$1<>\"\",{base_formula[1:]},\"\")")
                    cell.number_format = self.deliv_format
                
                # Apply borders
                for col in range(1, 22):
                    ws.cell(row=current_row, column=col).border = self.border
                
                detail_rows.append(current_row)
                current_row += 1
            
            # Add totals to group header
            if detail_rows:
                for col_idx in [7, 8, 12]:
                    col_letter = chr(64 + col_idx)
                    cell = ws.cell(row=group_start_row, column=col_idx,
                        value=f"=SUM({col_letter}{detail_rows[0]}:{col_letter}{detail_rows[-1]})")
                    cell.number_format = self.deliv_format
                
                for col_idx, col_letter in [(15, "O"), (17, "Q"), (19, "S"), (21, "U")]:
                    cell = ws.cell(row=group_start_row, column=col_idx,
                        value=f"=IF($M$1<>\"\",SUM({col_letter}{detail_rows[0]}:{col_letter}{detail_rows[-1]}),\"\")")
                    cell.number_format = self.deliv_format
            
            # Apply grouping (collapsible rows)
            for row_num in detail_rows:
                ws.row_dimensions[row_num].outline_level = 1
                ws.row_dimensions[row_num].hidden = False
        
        # Set column widths
        self._set_master_column_widths(ws)
        
        # Configure outline settings
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.summaryRight = False
    
    def _write_expiry_sheet(self, expiry_date: datetime, positions: List[Dict],
                           prices: Dict[str, float], prefix: str, position_type: str = ""):
        """Write sheet for specific expiry with Bloomberg formulas"""
        sheet_name = f"{prefix}_Expiry_{expiry_date.strftime('%d_%m_%Y')}"
        ws = self.wb.create_sheet(sheet_name)
        
        # Filter positions for this expiry
        expiry_positions = [p for p in positions if p['expiry'].date() == expiry_date.date()]
        
        if not expiry_positions:
            return
        
        # Headers matching the master sheet
        headers = [
            "Underlying", "Symbol", "Expiry", "Position", "Type", "Strike",
            "System Deliverable", "Override Deliverable", "System Price",
            "Override Price", "BBG Price", "BBG Deliverable",
            "", "System Price -%", "System Deliv -%", "System Price +%", "System Deliv +%",
            "BBG Price -%", "BBG Deliv -%", "BBG Price +%", "BBG Deliv +%"
        ]
        
        # Add sensitivity percentage input in M1
        ws.cell(row=1, column=13, value=10)  # Default 10% sensitivity
        ws.cell(row=1, column=13).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            if col != 13:  # Skip the separator column
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Group positions by underlying
        grouped = self._group_by_underlying(expiry_positions)
        
        current_row = 3
        
        for underlying in sorted(grouped.keys()):
            underlying_positions = grouped[underlying]
            group_start_row = current_row
            
            # Write group header
            cell = ws.cell(row=current_row, column=1, value=underlying)
            cell.font = self.group_font
            
            # Apply group formatting
            for col in range(1, 22):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = self.group_fill
                cell.border = self.border
            
            # Get spot price from Yahoo (try different keys)
            spot_price = None
            if prices:
                # Try underlying first
                spot_price = prices.get(underlying)
                # If not found, try the symbol from first position
                if not spot_price and underlying_positions:
                    first_symbol = underlying_positions[0].get('symbol', '')
                    spot_price = prices.get(first_symbol)
                # If still not found, check if underlying has a space and try first part
                if not spot_price and ' ' in underlying:
                    base_ticker = underlying.split(' ')[0]
                    spot_price = prices.get(base_ticker)
            
            # System price in group header (Yahoo price)
            if spot_price:
                ws.cell(row=current_row, column=9, value=spot_price).number_format = self.price_format
            else:
                ws.cell(row=current_row, column=9, value="").number_format = self.price_format
            
            # Override price (blank for user input)
            ws.cell(row=current_row, column=10, value="").number_format = self.price_format
            
            # Bloomberg price formula
            bbg_formula = f'=BDP(A{current_row},"PX_LAST")'
            ws.cell(row=current_row, column=11, value=bbg_formula).number_format = self.price_format
            
            # Sensitivity price formulas for group header
            ws.cell(row=current_row, column=14, 
                value=f"=IF($M$1<>\"\",I{current_row}*(1-$M$1/100),\"\")").number_format = self.price_format
            ws.cell(row=current_row, column=16,
                value=f"=IF($M$1<>\"\",I{current_row}*(1+$M$1/100),\"\")").number_format = self.price_format
            ws.cell(row=current_row, column=18,
                value=f"=IF($M$1<>\"\",K{current_row}*(1-$M$1/100),\"\")").number_format = self.price_format
            ws.cell(row=current_row, column=20,
                value=f"=IF($M$1<>\"\",K{current_row}*(1+$M$1/100),\"\")").number_format = self.price_format
            
            detail_rows = []
            current_row += 1
            
            # Write detail rows
            for pos in sorted(underlying_positions, key=lambda x: (x['expiry'], x['strike'])):
                ws.cell(row=current_row, column=2, value=pos['ticker'])
                ws.cell(row=current_row, column=3, value=pos['expiry'].strftime('%d/%m/%Y'))
                ws.cell(row=current_row, column=4, value=pos['lots'])
                ws.cell(row=current_row, column=5, value=pos['security_type'])
                
                if pos['strike'] > 0:
                    ws.cell(row=current_row, column=6, value=pos['strike']).number_format = self.price_format
                
                # Deliverable formulas
                for col_idx, price_col in [(7, "I"), (8, "J"), (12, "K")]:
                    formula = self._create_deliverable_formula(
                        current_row, group_start_row, pos['security_type'], 
                        pos['strike'], pos['lots'], price_col
                    )
                    ws.cell(row=current_row, column=col_idx, value=formula).number_format = self.deliv_format
                
                # Sensitivity deliverable formulas
                for col_idx, price_col in [(15, "N"), (17, "P"), (19, "R"), (21, "T")]:
                    base_formula = self._create_deliverable_formula(
                        current_row, group_start_row, pos['security_type'],
                        pos['strike'], pos['lots'], price_col
                    )
                    cell = ws.cell(row=current_row, column=col_idx, 
                        value=f"=IF($M$1<>\"\",{base_formula[1:]},\"\")")
                    cell.number_format = self.deliv_format
                
                # Apply borders
                for col in range(1, 22):
                    ws.cell(row=current_row, column=col).border = self.border
                
                detail_rows.append(current_row)
                current_row += 1
            
            # Add totals to group header
            if detail_rows:
                for col_idx in [7, 8, 12]:
                    col_letter = chr(64 + col_idx)
                    cell = ws.cell(row=group_start_row, column=col_idx,
                        value=f"=SUM({col_letter}{detail_rows[0]}:{col_letter}{detail_rows[-1]})")
                    cell.number_format = self.deliv_format
                
                for col_idx, col_letter in [(15, "O"), (17, "Q"), (19, "S"), (21, "U")]:
                    cell = ws.cell(row=group_start_row, column=col_idx,
                        value=f"=IF($M$1<>\"\",SUM({col_letter}{detail_rows[0]}:{col_letter}{detail_rows[-1]}),\"\")")
                    cell.number_format = self.deliv_format
            
            # Apply grouping (collapsible rows)
            for row_num in detail_rows:
                ws.row_dimensions[row_num].outline_level = 1
                ws.row_dimensions[row_num].hidden = False
        
        # Set column widths
        self._set_master_column_widths(ws)
        
        # Configure outline settings
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.summaryRight = False
    
    def _write_iv_master_sheet(self, positions: List[Dict], prices: Dict[str, float], prefix: str, position_type: str = ""):
        """Write IV sheet with all expiries"""
        sheet_name = f"{prefix}_IV_All_Expiries"
        ws = self.wb.create_sheet(sheet_name)
        
        # Grand total row
        grand_total_row = 1
        ws.cell(row=grand_total_row, column=1, value="GRAND TOTAL").font = Font(bold=True, size=12)
        
        for col in range(1, 17):
            cell = ws.cell(row=grand_total_row, column=col)
            cell.fill = self.grand_total_fill
            cell.border = self.border
        
        # Headers
        headers = [
            "Underlying", "Symbol", "Expiry", "Position", "Type", "Strike", "Lot Size",
            "System IV (INR)", "System IV (USD)", "Override IV (INR)", "Override IV (USD)",
            "System Price", "Override Price", "BBG Price", "BBG IV (INR)", "BBG IV (USD)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Group positions and write data
        grouped = self._group_by_underlying(positions)
        current_row = 3
        
        for underlying in sorted(grouped.keys()):
            underlying_positions = grouped[underlying]
            group_start_row = current_row
            
            # Group header
            ws.cell(row=current_row, column=1, value=underlying).font = self.group_font
            
            for col in range(1, 17):
                ws.cell(row=current_row, column=col).fill = self.group_fill
                ws.cell(row=current_row, column=col).border = self.border
            
            # Get Yahoo price (System Price)
            spot_price = None
            if prices:
                # Try underlying first
                spot_price = prices.get(underlying)
                # If not found, try the symbol from first position
                if not spot_price and underlying_positions:
                    first_symbol = underlying_positions[0].get('symbol', '')
                    spot_price = prices.get(first_symbol)
                # If still not found, check if underlying has a space and try first part
                if not spot_price and ' ' in underlying:
                    base_ticker = underlying.split(' ')[0]
                    spot_price = prices.get(base_ticker)
            
            # System price (Yahoo price)
            if spot_price:
                ws.cell(row=current_row, column=12, value=spot_price).number_format = self.price_format
            else:
                ws.cell(row=current_row, column=12, value="").number_format = self.price_format
            
            ws.cell(row=current_row, column=13, value="").number_format = self.price_format
            ws.cell(row=current_row, column=14, value=f'=BDP(A{current_row},"PX_LAST")').number_format = self.price_format
            
            detail_rows = []
            current_row += 1
            
            # Write detail rows with IV formulas
            for pos in sorted(underlying_positions, key=lambda x: (x['expiry'], x['strike'])):
                ws.cell(row=current_row, column=2, value=pos['ticker'])
                ws.cell(row=current_row, column=3, value=pos['expiry'].strftime('%d/%m/%Y'))
                ws.cell(row=current_row, column=4, value=pos['lots'])
                ws.cell(row=current_row, column=5, value=pos['security_type'])
                
                if pos['strike'] > 0:
                    ws.cell(row=current_row, column=6, value=pos['strike']).number_format = self.price_format
                
                ws.cell(row=current_row, column=7, value=pos['lot_size'])
                
                # IV formulas
                for col_idx, price_col in [(8, "L"), (10, "M"), (15, "N")]:
                    formula = self._create_iv_formula(current_row, group_start_row, price_col)
                    ws.cell(row=current_row, column=col_idx, value=formula).number_format = self.iv_format
                
                # USD conversions
                for inr_col, usd_col in [(8, 9), (10, 11), (15, 16)]:
                    inr_cell = chr(64 + inr_col)
                    ws.cell(row=current_row, column=usd_col, 
                        value=f"={inr_cell}{current_row}/{self.usdinr_rate}").number_format = self.iv_format
                
                for col in range(1, 17):
                    ws.cell(row=current_row, column=col).border = self.border
                
                detail_rows.append(current_row)
                current_row += 1
            
            # Apply grouping
            for row_num in detail_rows:
                ws.row_dimensions[row_num].outline_level = 1
                ws.row_dimensions[row_num].hidden = False
        
        # Grand totals
        for col_idx, col_letter in [(8, "H"), (9, "I"), (10, "J"), (11, "K"), (15, "O"), (16, "P")]:
            ws.cell(row=grand_total_row, column=col_idx, 
                value=f"=SUM({col_letter}3:{col_letter}1000)").number_format = self.iv_format
        
        self._set_iv_column_widths(ws)
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.summaryRight = False
    
    def _write_comparison_sheet(self, pre_positions: List[Dict], post_positions: List[Dict], 
                               prices: Dict[str, float]):
        """Write comparison sheet between pre and post trade positions"""
        ws = self.wb.create_sheet("Deliverables_Comparison")
        
        ws.cell(row=1, column=1, value="DELIVERABLES COMPARISON").font = Font(bold=True, size=12)
        
        # Headers
        headers = [
            "Ticker", "Symbol", "Type", "Strike", 
            "Pre-Trade Deliv", "Post-Trade Deliv", "Deliv Change",
            "Pre-Trade IV", "Post-Trade IV", "IV Change"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Create comparison data
        pre_dict = {p['ticker']: p for p in pre_positions}
        post_dict = {p['ticker']: p for p in post_positions}
        all_tickers = set(pre_dict.keys()) | set(post_dict.keys())
        
        row = 4
        for ticker in sorted(all_tickers):
            pre = pre_dict.get(ticker)
            post = post_dict.get(ticker)
            
            ws.cell(row=row, column=1, value=ticker)
            
            if pre or post:
                ref = pre if pre else post
                ws.cell(row=row, column=2, value=ref['symbol'])
                ws.cell(row=row, column=3, value=ref['security_type'])
                if ref['strike'] > 0:
                    ws.cell(row=row, column=4, value=ref['strike']).number_format = self.price_format
            
            # Calculate deliverables
            pre_deliv = self._calculate_position_deliverable(pre, prices) if pre else 0
            post_deliv = self._calculate_position_deliverable(post, prices) if post else 0
            
            ws.cell(row=row, column=5, value=pre_deliv)
            ws.cell(row=row, column=6, value=post_deliv)
            ws.cell(row=row, column=7, value=post_deliv - pre_deliv)
            
            # Color code changes
            change_cell = ws.cell(row=row, column=7)
            if post_deliv - pre_deliv > 0:
                change_cell.fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
            elif post_deliv - pre_deliv < 0:
                change_cell.fill = PatternFill(start_color="E5FFE5", end_color="E5FFE5", fill_type="solid")
            
            # Apply borders
            for col in range(1, 11):
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Set column widths
        for col, width in [('A', 35), ('B', 15), ('C', 10), ('D', 10), ('E', 15), 
                          ('F', 15), ('G', 15), ('H', 15), ('I', 15), ('J', 15)]:
            ws.column_dimensions[col].width = width
    
    def _write_all_positions_sheet(self, positions: List[Dict], sheet_name: str):
        """Write sheet with all positions as simple list with Yahoo prices"""
        if not positions:
            return
            
        ws = self.wb.create_sheet(sheet_name)
        
        # Pass prices to this method
        prices = self.prices if hasattr(self, 'prices') else {}
        
        headers = ["Underlying", "Symbol", "Expiry", "Position", "Type", "Strike", "Lot Size", 
                  "Yahoo Price", "Moneyness", "Deliverable"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        sorted_positions = sorted(positions, key=lambda x: (x['underlying'], x['expiry'], x['strike']))
        
        current_row = 2
        for pos in sorted_positions:
            ws.cell(row=current_row, column=1, value=pos['underlying'])
            ws.cell(row=current_row, column=2, value=pos['ticker'])
            ws.cell(row=current_row, column=3, value=pos['expiry'].strftime('%d/%m/%Y'))
            ws.cell(row=current_row, column=4, value=pos['lots'])
            ws.cell(row=current_row, column=5, value=pos['security_type'])
            
            if pos['strike'] > 0:
                ws.cell(row=current_row, column=6, value=pos['strike']).number_format = self.price_format
            
            ws.cell(row=current_row, column=7, value=pos['lot_size'])
            
            # Get Yahoo price
            spot_price = None
            if prices:
                # Try underlying first
                spot_price = prices.get(pos['underlying'])
                # If not found, try symbol
                if not spot_price:
                    spot_price = prices.get(pos['symbol'])
                # If still not found and underlying has space, try base ticker
                if not spot_price and ' ' in pos['underlying']:
                    base_ticker = pos['underlying'].split(' ')[0]
                    spot_price = prices.get(base_ticker)
            
            # Yahoo Price column
            if spot_price:
                ws.cell(row=current_row, column=8, value=spot_price).number_format = self.price_format
            else:
                ws.cell(row=current_row, column=8, value="N/A")
            
            # Moneyness column
            moneyness = ""
            if spot_price and pos['security_type'] in ['Call', 'Put']:
                if pos['security_type'] == 'Call':
                    if spot_price > pos['strike']:
                        moneyness = "ITM"
                    elif spot_price < pos['strike']:
                        moneyness = "OTM"
                    else:
                        moneyness = "ATM"
                else:  # Put
                    if spot_price < pos['strike']:
                        moneyness = "ITM"
                    elif spot_price > pos['strike']:
                        moneyness = "OTM"
                    else:
                        moneyness = "ATM"
            elif pos['security_type'] == 'Futures':
                moneyness = "N/A"
            ws.cell(row=current_row, column=9, value=moneyness)
            
            # Deliverable column
            deliverable = 0
            if spot_price or pos['security_type'] == 'Futures':
                deliverable = self._calculate_position_deliverable(pos, prices)
            ws.cell(row=current_row, column=10, value=deliverable).number_format = self.deliv_format
            
            # Apply borders
            for col in range(1, 11):
                ws.cell(row=current_row, column=col).border = self.border
            
            current_row += 1
        
        # Add totals row
        total_row = current_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{current_row-1})").font = Font(bold=True)
        ws.cell(row=total_row, column=10, value=f"=SUM(J2:J{current_row-1})").font = Font(bold=True)
        ws.cell(row=total_row, column=10).number_format = self.deliv_format
        
        for col in range(1, 11):
            ws.cell(row=total_row, column=col).border = self.border
            ws.cell(row=total_row, column=col).fill = self.header_fill
        
        # Set column widths
        widths = {'A': 25, 'B': 35, 'C': 12, 'D': 10, 'E': 8, 'F': 10, 'G': 10, 
                 'H': 12, 'I': 12, 'J': 12}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    
    # Helper methods
    def _group_by_underlying(self, positions: List[Dict]) -> Dict[str, List[Dict]]:
        """Group positions by underlying"""
        grouped = {}
        for pos in positions:
            underlying = pos['underlying']
            if underlying not in grouped:
                grouped[underlying] = []
            grouped[underlying].append(pos)
        return grouped
    
    def _calculate_position_deliverable(self, pos: Dict, prices: Dict[str, float]) -> float:
        """Calculate deliverable for a position"""
        if not pos:
            return 0
        
        # Try to get price using multiple keys
        spot_price = 0
        if prices:
            # Try underlying first
            spot_price = prices.get(pos['underlying'], 0)
            # If not found, try symbol
            if not spot_price:
                spot_price = prices.get(pos['symbol'], 0)
            # If still not found and underlying has space, try base ticker
            if not spot_price and ' ' in pos['underlying']:
                base_ticker = pos['underlying'].split(' ')[0]
                spot_price = prices.get(base_ticker, 0)
        
        if pos['security_type'] == 'Futures':
            return pos['lots']
        elif pos['security_type'] == 'Call':
            if spot_price > 0 and spot_price > pos['strike']:
                return pos['lots']
            return 0
        elif pos['security_type'] == 'Put':
            if spot_price > 0 and spot_price < pos['strike']:
                return -pos['lots']
            return 0
        
        return 0
    
    def _create_deliverable_formula(self, row: int, group_header_row: int, 
                                   security_type: str, strike: float, lots: float, 
                                   price_column: str) -> str:
        """Create Excel formula for deliverable calculation"""
        type_cell = f"E{row}"
        position_cell = f"D{row}"
        strike_cell = f"F{row}"
        price_cell = f"${price_column}${group_header_row}"
        
        formula = (
            f'=IF({type_cell}="Futures",{position_cell},'
            f'IF({type_cell}="Call",IF({price_cell}>{strike_cell},{position_cell},0),'
            f'IF({type_cell}="Put",IF({price_cell}<{strike_cell},-{position_cell},0),0)))'
        )
        
        return formula
    
    def _create_iv_formula(self, row: int, group_header_row: int, price_column: str) -> str:
        """Create Excel formula for IV calculation"""
        type_cell = f"E{row}"
        position_cell = f"D{row}"
        strike_cell = f"F{row}"
        lot_size_cell = f"G{row}"
        price_cell = f"${price_column}${group_header_row}"
        
        formula = (
            f'=IF({type_cell}="Futures",0,'
            f'IF({type_cell}="Call",IF({price_cell}>{strike_cell},{position_cell}*{lot_size_cell}*({price_cell}-{strike_cell}),0),'
            f'IF({type_cell}="Put",IF({price_cell}<{strike_cell},{position_cell}*{lot_size_cell}*({strike_cell}-{price_cell}),0),0)))'
        )
        
        return formula
    
    def _set_master_column_widths(self, ws: Worksheet):
        """Set column widths for master sheet"""
        widths = {
            'A': 25, 'B': 35, 'C': 12, 'D': 10, 'E': 8, 'F': 10,
            'G': 15, 'H': 15, 'I': 12, 'J': 12, 'K': 12, 'L': 15,
            'M': 8, 'N': 14, 'O': 14, 'P': 14, 'Q': 14, 'R': 14,
            'S': 14, 'T': 14, 'U': 14
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    
    def _set_iv_column_widths(self, ws: Worksheet):
        """Set column widths for IV sheets"""
        widths = {
            'A': 25, 'B': 35, 'C': 12, 'D': 10, 'E': 8, 'F': 10, 'G': 10,
            'H': 15, 'I': 15, 'J': 15, 'K': 15, 'L': 12, 'M': 12, 'N': 12,
            'O': 15, 'P': 15
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
