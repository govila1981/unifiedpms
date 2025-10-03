"""
Streamlit Web App for Stock Price Fetcher
Deploy this on Streamlit Cloud for easy access
"""

import streamlit as st
import yfinance as yf
import pandas as pd
from datetime import datetime
import time
import io
import os

# Page config
st.set_page_config(
    page_title="Stock Price Fetcher",
    page_icon="📈",
    layout="wide"
)

# Title
st.title("📈 Stock Price Fetcher")
st.markdown("### Fetch live prices from Yahoo Finance for Indian stocks")

# Sidebar
st.sidebar.header("⚙️ Settings")
delay = st.sidebar.slider("Delay between requests (seconds)", 0.1, 1.0, 0.2, 0.1)
exchange_default = st.sidebar.selectbox("Default Exchange", ["NSE (.NS)", "BSE (.BO)"])

st.sidebar.markdown("---")
st.sidebar.markdown("### 📖 How to use")
st.sidebar.markdown("""
1. Upload your CSV/Excel file
2. Select the symbol column
3. Click 'Fetch Prices'
4. Download the updated file
""")

# File selection option
st.markdown("---")
file_option = st.radio(
    "📁 Choose data source:",
    ["Use default file (Nifty 500 stocks)", "Upload my own file"],
    horizontal=True
)

uploaded_file = None
df = None

if file_option == "Upload my own file":
    uploaded_file = st.file_uploader(
        "Upload your file (CSV or Excel)",
        type=['csv', 'xlsx', 'xls'],
        help="File should contain stock symbols"
    )
else:
    # Use default file
    try:
        default_file = "default_stocks.csv"
        if os.path.exists(default_file):
            df = pd.read_csv(default_file)
            st.success(f"✓ Using default file: {default_file}")
            st.info(f"Found {len(df)} rows and {len(df.columns)} columns")
        else:
            st.warning(f"⚠️ Default file '{default_file}' not found. Please upload a file instead.")
            file_option = "Upload my own file"
    except Exception as e:
        st.error(f"Error loading default file: {str(e)}")
        file_option = "Upload my own file"

if uploaded_file:
    # Read the uploaded file
    try:
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file)

        st.success(f"✓ File loaded: {uploaded_file.name}")
        st.info(f"Found {len(df)} rows and {len(df.columns)} columns")
    except Exception as e:
        st.error(f"Error reading file: {str(e)}")
        st.exception(e)

if df is not None:
    # Process the dataframe (whether from default file or uploaded file)
    try:

        # Show preview
        with st.expander("📋 Preview Data (first 5 rows)"):
            st.dataframe(df.head())

        # Column selection
        st.markdown("---")
        col1, col2 = st.columns([2, 1])

        with col1:
            st.subheader("🎯 Select Columns")

            # Auto-detect columns
            def detect_yahoo_column(df):
                yahoo_patterns = ['yahoo', 'ticker', 'yahoo_ticker', 'yahoo ticker']
                for col in df.columns:
                    col_lower = col.lower().strip()
                    for pattern in yahoo_patterns:
                        if pattern in col_lower:
                            sample_value = str(df[col].iloc[0]) if len(df) > 0 else ''
                            if '.NS' in sample_value or '.BO' in sample_value:
                                return col
                return None

            def detect_symbol_column(df):
                symbol_patterns = ['symbol', 'ticker', 'stock', 'code', 'scrip']
                for col in df.columns:
                    col_lower = col.lower().strip()
                    if col_lower in symbol_patterns:
                        return col
                for col in df.columns:
                    col_lower = col.lower().strip()
                    for pattern in symbol_patterns:
                        if pattern in col_lower:
                            return col
                return df.columns[0]  # Default to first column

            def detect_exchange_column(df):
                exchange_patterns = ['exchange', 'market', 'series']
                for col in df.columns:
                    col_lower = col.lower().strip()
                    for pattern in exchange_patterns:
                        if pattern in col_lower:
                            return col
                return None

            # Detect columns
            yahoo_col_detected = detect_yahoo_column(df)
            symbol_col_detected = detect_symbol_column(df)
            exchange_col_detected = detect_exchange_column(df)

            # Let user select or confirm
            if yahoo_col_detected:
                st.success(f"✓ Found Yahoo Ticker column: **{yahoo_col_detected}**")
                yahoo_col = st.selectbox("Yahoo Ticker Column", df.columns,
                                        index=df.columns.get_loc(yahoo_col_detected))
                symbol_col = None
                exchange_col = None
            else:
                st.info("Yahoo Ticker column not found. Will create tickers from symbol.")
                default_idx = df.columns.get_loc(symbol_col_detected) if symbol_col_detected else 0
                symbol_col = st.selectbox("Symbol Column", df.columns, index=default_idx)

                yahoo_col = None

                if exchange_col_detected:
                    exchange_col = st.selectbox("Exchange Column (optional)",
                                               ['None'] + list(df.columns),
                                               index=df.columns.get_loc(exchange_col_detected) + 1)
                    if exchange_col == 'None':
                        exchange_col = None
                else:
                    exchange_col = st.selectbox("Exchange Column (optional)",
                                               ['None'] + list(df.columns))
                    if exchange_col == 'None':
                        exchange_col = None

        with col2:
            st.subheader("📊 Statistics")
            st.metric("Total Rows", len(df))
            st.metric("Columns", len(df.columns))
            if symbol_col:
                unique_symbols = df[symbol_col].nunique()
                st.metric("Unique Symbols", unique_symbols)

        # Fetch button
        st.markdown("---")

        if st.button("🚀 Fetch Prices", type="primary", use_container_width=True):

            # Create Yahoo tickers if needed
            def create_yahoo_ticker(symbol, exchange=None):
                if pd.isna(symbol) or symbol == '':
                    return None

                symbol = str(symbol).strip()

                if '.NS' in symbol or '.BO' in symbol:
                    return symbol

                suffix = '.NS' if 'NSE' in exchange_default else '.BO'

                if exchange:
                    exchange_str = str(exchange).upper()
                    if 'BSE' in exchange_str or 'B' == exchange_str:
                        suffix = '.BO'
                    elif 'NSE' in exchange_str or 'N' == exchange_str:
                        suffix = '.NS'

                return f"{symbol}{suffix}"

            # Prepare tickers
            if yahoo_col:
                df['Yahoo_Ticker'] = df[yahoo_col]
            else:
                if exchange_col:
                    df['Yahoo_Ticker'] = df.apply(
                        lambda row: create_yahoo_ticker(row[symbol_col], row[exchange_col]),
                        axis=1
                    )
                else:
                    df['Yahoo_Ticker'] = df[symbol_col].apply(
                        lambda x: create_yahoo_ticker(x, None)
                    )

            # Filter valid tickers
            valid_df = df[df['Yahoo_Ticker'].notna()].copy()

            st.info(f"Fetching prices for {len(valid_df)} stocks...")

            # Progress bar
            progress_bar = st.progress(0)
            status_text = st.empty()

            prices = []
            timestamps = []

            for idx, row in valid_df.iterrows():
                ticker = row['Yahoo_Ticker']

                try:
                    stock = yf.Ticker(ticker)
                    info = stock.info
                    current_price = info.get('currentPrice') or info.get('regularMarketPrice')

                    if current_price:
                        prices.append(round(current_price, 2))
                    else:
                        prices.append('N/A')
                except:
                    prices.append('Error')

                timestamps.append(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

                # Update progress
                progress = (len(prices)) / len(valid_df)
                progress_bar.progress(progress)
                status_text.text(f"Fetching... {len(prices)}/{len(valid_df)} stocks")

                # Delay
                time.sleep(delay)

            # Add results to dataframe
            valid_df['Current_Price_INR'] = prices
            valid_df['Last_Updated'] = timestamps

            progress_bar.progress(1.0)
            status_text.text("✓ Complete!")

            # Display results
            st.success(f"✓ Successfully fetched prices for {len(valid_df)} stocks!")

            # Statistics
            col1, col2, col3 = st.columns(3)
            successful = sum(1 for p in prices if isinstance(p, (int, float)))
            failed = sum(1 for p in prices if p in ['N/A', 'Error'])

            with col1:
                st.metric("Successful", successful, delta=f"{successful/len(prices)*100:.1f}%")
            with col2:
                st.metric("Failed", failed)
            with col3:
                avg_price = sum(p for p in prices if isinstance(p, (int, float))) / max(successful, 1)
                st.metric("Avg Price", f"₹{avg_price:.2f}")

            # Show results
            st.markdown("---")
            st.subheader("📋 Results")

            # Display options
            display_cols = st.multiselect(
                "Select columns to display",
                valid_df.columns.tolist(),
                default=[col for col in valid_df.columns if col in [symbol_col, 'Company Name', 'Yahoo_Ticker', 'Current_Price_INR', 'Last_Updated']][:5]
            )

            if display_cols:
                st.dataframe(valid_df[display_cols], use_container_width=True)
            else:
                st.dataframe(valid_df, use_container_width=True)

            # Download button
            st.markdown("---")

            # Convert to Excel in memory
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                valid_df.to_excel(writer, sheet_name='Prices', index=False)

                # Get the worksheet
                worksheet = writer.sheets['Prices']

                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter

                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Apply light blue background to Current_Price_INR and Bloomberg Code columns
                from openpyxl.styles import PatternFill
                light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

                # Find column indices
                headers = [cell.value for cell in worksheet[1]]
                price_col_idx = None
                bloomberg_col_idx = None

                for idx, header in enumerate(headers, 1):
                    if header == 'Current_Price_INR':
                        price_col_idx = idx
                    elif header == 'Bloomberg Code':
                        bloomberg_col_idx = idx

                # Apply blue fill to the columns
                if price_col_idx:
                    col_letter = worksheet.cell(1, price_col_idx).column_letter
                    for row in range(1, len(valid_df) + 2):
                        worksheet[f'{col_letter}{row}'].fill = light_blue_fill

                if bloomberg_col_idx:
                    col_letter = worksheet.cell(1, bloomberg_col_idx).column_letter
                    for row in range(1, len(valid_df) + 2):
                        worksheet[f'{col_letter}{row}'].fill = light_blue_fill

            excel_data = output.getvalue()

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"Prices_Updated_{timestamp}.xlsx"

            st.download_button(
                label="⬇️ Download Excel File",
                data=excel_data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )

            # Store in session state for future downloads
            st.session_state['result_df'] = valid_df
            st.session_state['excel_data'] = excel_data
            st.session_state['filename'] = filename

    except Exception as e:
        st.error(f"Error processing file: {str(e)}")
        st.exception(e)

else:
    # Show instructions when no file selected
    st.info("👆 Select a data source to get started!")

    st.markdown("---")
    st.markdown("### 📝 Sample File Format")
    st.markdown("Your file should contain at least one column with stock symbols:")

    sample_df = pd.DataFrame({
        'Symbol': ['RELIANCE', 'TCS', 'INFY', 'HDFCBANK', 'ICICIBANK'],
        'Company Name': ['Reliance Industries', 'Tata Consultancy Services', 'Infosys', 'HDFC Bank', 'ICICI Bank'],
        'Exchange': ['NSE', 'NSE', 'NSE', 'NSE', 'NSE']
    })

    st.dataframe(sample_df, use_container_width=True)

    st.markdown("---")
    st.markdown("### ✨ Features")
    col1, col2 = st.columns(2)

    with col1:
        st.markdown("""
        - 📊 Auto-detects symbol columns
        - 🎯 Creates Yahoo Finance tickers
        - 📈 Fetches live prices
        - 💾 Downloads as Excel
        """)

    with col2:
        st.markdown("""
        - 🔄 Handles NSE/BSE exchanges
        - ⚡ Configurable delays
        - 📋 Preserves original data
        - 🎨 Clean interface
        """)

# Footer
st.markdown("---")
st.markdown(
    "<div style='text-align: center; color: gray;'>"
    "Made with ❤️ using Streamlit | Data from Yahoo Finance"
    "</div>",
    unsafe_allow_html=True
)