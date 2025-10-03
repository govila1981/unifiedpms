"""
Streamlit Futures Delivery Calculator with Position Reconciliation
Web application for calculating physical delivery and reconciling positions
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import io
import tempfile
import os
import logging
from typing import Dict, List, Optional
import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from copy import copy

# Import your modules
from input_parser import InputParser, Position
from price_fetcher import PriceFetcher
from excel_writer import ExcelWriter
from recon_module import PositionReconciliation

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Page config
st.set_page_config(
    page_title="Futures Delivery Calculator",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        padding: 1rem;
        border-bottom: 3px solid #1f77b4;
        margin-bottom: 2rem;
    }
    .sub-header {
        font-size: 1.5rem;
        color: #333;
        margin-top: 1.5rem;
        margin-bottom: 1rem;
    }
    .info-box {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
    }
    .success-box {
        background-color: #d4edda;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
        border: 1px solid #c3e6cb;
    }
    .recon-box {
        background-color: #e7f3ff;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
        border: 1px solid #b3d9ff;
    }
</style>
""", unsafe_allow_html=True)


class StreamlitDeliveryApp:
    """Main Streamlit application class"""
    
    def __init__(self):
        self.initialize_session_state()
        self.recon_module = PositionReconciliation()
    
    def initialize_session_state(self):
        """Initialize session state variables"""
        if 'positions' not in st.session_state:
            st.session_state.positions = []
        if 'prices' not in st.session_state:
            st.session_state.prices = {}
        if 'unmapped_symbols' not in st.session_state:
            st.session_state.unmapped_symbols = []
        if 'report_generated' not in st.session_state:
            st.session_state.report_generated = False
        if 'output_file' not in st.session_state:
            st.session_state.output_file = None
        if 'recon_results' not in st.session_state:
            st.session_state.recon_results = None
        if 'recon_file' not in st.session_state:
            st.session_state.recon_file = None
        if 'file_prefix' not in st.session_state:
            st.session_state.file_prefix = 'DELIVERY'
    
    def run(self):
        """Main application entry point"""
        # Header
        st.markdown('<h1 class="main-header">📊 Futures & Options Delivery Calculator</h1>', 
                   unsafe_allow_html=True)
        
        # Sidebar for configuration
        with st.sidebar:
            st.header("⚙️ Configuration")
            
            # USDINR Rate
            usdinr_rate = st.number_input(
                "USD/INR Exchange Rate",
                min_value=50.0,
                max_value=150.0,
                value=88.0,
                step=0.1,
                help="Current USD to INR exchange rate for IV calculations"
            )
            
            # Mapping file upload
            st.subheader("📁 Symbol Mapping File")
            mapping_file = st.file_uploader(
                "Upload futures mapping CSV",
                type=['csv'],
                help="CSV file with symbol to ticker mappings"
            )
            
            mapping_file_path = None
            if not mapping_file:
                st.info("ℹ️ Using default 'futures mapping.csv'")
                # Try to find the mapping file
                possible_paths = ['futures mapping.csv', 'futures_mapping.csv']
                for path in possible_paths:
                    if os.path.exists(path):
                        mapping_file_path = path
                        break
                if not mapping_file_path:
                    mapping_file_path = 'futures mapping.csv'
            else:
                # Save uploaded mapping file temporarily
                with tempfile.NamedTemporaryFile(delete=False, suffix='.csv', mode='wb') as tmp_file:
                    tmp_file.write(mapping_file.getvalue())
                    mapping_file_path = tmp_file.name
            
            st.divider()
            
            # Price fetching options
            st.subheader("💹 Price Options")
            fetch_prices = st.checkbox("Fetch prices from Yahoo Finance", value=True)
            
            st.divider()
            
            # Reconciliation options
            st.subheader("🔄 Reconciliation (Optional)")
            st.info("Upload a recon file to compare positions")
            recon_file = st.file_uploader(
                "Upload reconciliation file",
                type=['xlsx', 'xls', 'csv'],
                help="File with Symbol and Position columns to reconcile against",
                key="recon_uploader"
            )
            
            if recon_file:
                st.success(f"✅ Recon file loaded: {recon_file.name}")
                st.session_state.recon_file = recon_file
        
        # Main content area with tabs
        tabs = st.tabs(["📤 Upload & Process", "📊 Positions Review", 
                        "💰 Deliverables Preview", "🔄 Reconciliation", "📥 Download Reports"])
        
        with tabs[0]:
            self.upload_and_process_tab(mapping_file_path, usdinr_rate, fetch_prices)
        
        with tabs[1]:
            self.positions_review_tab()
        
        with tabs[2]:
            self.deliverables_preview_tab()
        
        with tabs[3]:
            self.reconciliation_tab()
        
        with tabs[4]:
            self.download_reports_tab()
    
    def upload_and_process_tab(self, mapping_file_path, usdinr_rate, fetch_prices):
        """Handle file upload and processing"""
        st.markdown('<h2 class="sub-header">Upload Position File</h2>', unsafe_allow_html=True)
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            uploaded_file = st.file_uploader(
                "Choose your position file",
                type=['xlsx', 'xls', 'csv'],
                help="Upload BOD, CONTRACT, or MS format position file"
            )
        
        with col2:
            if uploaded_file:
                st.markdown('<div class="info-box">', unsafe_allow_html=True)
                st.write("**File Details:**")
                st.write(f"📁 Name: {uploaded_file.name}")
                st.write(f"📏 Size: {uploaded_file.size:,} bytes")
                st.markdown('</div>', unsafe_allow_html=True)
        
        if uploaded_file and mapping_file_path:
            # Process button
            if st.button("🚀 Process File", type="primary", use_container_width=True):
                with st.spinner("Processing position file..."):
                    success, message = self.process_file(
                        uploaded_file, mapping_file_path, None, usdinr_rate, fetch_prices
                    )
                    
                    if success:
                        st.success(f"✅ {message}")
                        
                        # If recon file is uploaded, perform reconciliation automatically
                        if st.session_state.recon_file:
                            self.perform_reconciliation()
                        
                        st.balloons()
                    else:
                        st.error(f"❌ {message}")
    
    def process_file(self, uploaded_file, mapping_file_path, password, usdinr_rate, fetch_prices):
        """Process the uploaded file"""
        try:
            # Save uploaded file temporarily
            suffix = os.path.splitext(uploaded_file.name)[1]
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix, mode='wb') as tmp_file:
                tmp_file.write(uploaded_file.getvalue())
                input_file_path = tmp_file.name
            
            # Parse positions
            parser = InputParser(mapping_file_path)
            positions = parser.parse_file(input_file_path)
            
            if not positions:
                return False, "No valid positions found in the file"
            
            st.session_state.positions = positions
            st.session_state.unmapped_symbols = parser.unmapped_symbols
            
            # Fetch prices if enabled
            if fetch_prices:
                with st.spinner("Fetching prices from Yahoo Finance..."):
                    price_fetcher = PriceFetcher()
                    symbols_to_fetch = list(set(p.symbol for p in positions))
                    symbol_prices = price_fetcher.fetch_prices_for_symbols(symbols_to_fetch)
                    
                    # Map to underlying tickers
                    symbol_map = {}
                    for p in positions:
                        symbol_map[p.underlying_ticker] = p.symbol
                    
                    prices = {}
                    for underlying, symbol in symbol_map.items():
                        if symbol in symbol_prices:
                            prices[underlying] = symbol_prices[symbol]
                    
                    st.session_state.prices = prices
            
            # Generate Excel report
            with st.spinner("Generating Excel report..."):
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                format_type = getattr(parser, 'format_type', 'UNKNOWN')
                
                if format_type in ['BOD', 'CONTRACT']:
                    prefix = "GS_AURIGIN"
                    st.session_state.file_prefix = prefix
                elif format_type == 'MS':
                    prefix = "MS_WAFRA"
                    st.session_state.file_prefix = prefix
                else:
                    prefix = "DELIVERY"
                    st.session_state.file_prefix = prefix
                
                output_file = f"{prefix}_DELIVERY_{timestamp}.xlsx"
                
                writer = ExcelWriter(output_file, usdinr_rate)
                writer.create_report(positions, st.session_state.prices, parser.unmapped_symbols)
                
                st.session_state.output_file = output_file
                st.session_state.report_generated = True
            
            # Clean up temp file
            try:
                os.unlink(input_file_path)
            except:
                pass
            
            return True, f"Successfully processed {len(positions)} positions"
            
        except Exception as e:
            logger.error(f"Error processing file: {str(e)}")
            return False, f"Error processing file: {str(e)}"
    
    def perform_reconciliation(self):
        """Perform reconciliation if recon file is uploaded"""
        if not st.session_state.output_file or not st.session_state.recon_file:
            return
        
        try:
            # Save recon file temporarily
            suffix = os.path.splitext(st.session_state.recon_file.name)[1]
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix, mode='wb') as tmp_file:
                tmp_file.write(st.session_state.recon_file.getvalue())
                recon_file_path = tmp_file.name
            
            # Generate recon output filename with appropriate prefix
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            prefix = getattr(st.session_state, 'file_prefix', 'DELIVERY')
            recon_output_file = f"{prefix}_RECONCILIATION_{timestamp}.xlsx"
            
            # Perform reconciliation
            results = self.recon_module.perform_reconciliation(
                st.session_state.output_file,
                recon_file_path,
                recon_output_file
            )
            
            st.session_state.recon_results = results
            st.session_state.recon_output_file = recon_output_file
            
            # Clean up temp file
            try:
                os.unlink(recon_file_path)
            except:
                pass
                
        except Exception as e:
            logger.error(f"Error during reconciliation: {str(e)}")
            st.error(f"Reconciliation failed: {str(e)}")
    
    def generate_consolidated_report(self, delivery_file: str, recon_file: str) -> str:
        """
        Combine delivery report and reconciliation report into a single Excel file
        """
        try:
            # Generate output filename with appropriate prefix
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            prefix = getattr(st.session_state, 'file_prefix', 'DELIVERY')
            consolidated_file = f"{prefix}_CONSOLIDATED_{timestamp}.xlsx"
            
            # Load the delivery report as base
            wb_delivery = load_workbook(delivery_file)
            
            # Load the reconciliation report
            wb_recon = load_workbook(recon_file)
            
            # Copy all sheets from reconciliation report to delivery report
            # Prefix recon sheets with "RECON_" to distinguish them
            for sheet_name in wb_recon.sheetnames:
                source_sheet = wb_recon[sheet_name]
                
                # Create new sheet name with prefix
                if sheet_name == "Summary":
                    new_sheet_name = "RECON_Summary"
                else:
                    new_sheet_name = f"RECON_{sheet_name}" if not sheet_name.startswith("RECON_") else sheet_name
                
                # Ensure sheet name doesn't exceed Excel's 31 character limit
                if len(new_sheet_name) > 31:
                    new_sheet_name = new_sheet_name[:31]
                
                # Create new sheet in delivery workbook
                target_sheet = wb_delivery.create_sheet(new_sheet_name)
                
                # Copy all cells
                for row in source_sheet.iter_rows():
                    for cell in row:
                        target_cell = target_sheet.cell(
                            row=cell.row, 
                            column=cell.column, 
                            value=cell.value
                        )
                        
                        # Copy cell formatting
                        if cell.has_style:
                            target_cell.font = copy(cell.font)
                            target_cell.fill = copy(cell.fill)
                            target_cell.border = copy(cell.border)
                            target_cell.alignment = copy(cell.alignment)
                            target_cell.number_format = cell.number_format
                
                # Copy column widths
                for col_letter in source_sheet.column_dimensions:
                    target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width
                
                # Copy row heights
                for row_num in source_sheet.row_dimensions:
                    target_sheet.row_dimensions[row_num].height = source_sheet.row_dimensions[row_num].height
            
            # Add a consolidated summary sheet at the beginning
            summary_sheet = wb_delivery.create_sheet("CONSOLIDATED_SUMMARY", 0)
            
            # Add headers and summary information
            summary_sheet.cell(row=1, column=1, value="CONSOLIDATED DELIVERY & RECONCILIATION REPORT")
            summary_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
            
            summary_sheet.cell(row=3, column=1, value="Report Generated:")
            summary_sheet.cell(row=3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            
            summary_sheet.cell(row=5, column=1, value="CONTENTS:")
            summary_sheet.cell(row=5, column=1).font = Font(bold=True, size=12)
            
            # List delivery sheets
            summary_sheet.cell(row=7, column=1, value="DELIVERY REPORT SHEETS:")
            summary_sheet.cell(row=7, column=1).font = Font(bold=True)
            
            row = 8
            for sheet_name in wb_delivery.sheetnames:
                if not sheet_name.startswith("RECON_") and sheet_name != "CONSOLIDATED_SUMMARY":
                    summary_sheet.cell(row=row, column=2, value=f"• {sheet_name}")
                    row += 1
            
            # List reconciliation sheets
            row += 1
            summary_sheet.cell(row=row, column=1, value="RECONCILIATION REPORT SHEETS:")
            summary_sheet.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            
            for sheet_name in wb_delivery.sheetnames:
                if sheet_name.startswith("RECON_"):
                    display_name = sheet_name.replace("RECON_", "")
                    summary_sheet.cell(row=row, column=2, value=f"• {display_name}")
                    row += 1
            
            # Add reconciliation results if available
            if hasattr(st.session_state, 'recon_results') and st.session_state.recon_results:
                results = st.session_state.recon_results
                summary = results['summary']
                
                row += 2
                summary_sheet.cell(row=row, column=1, value="RECONCILIATION SUMMARY:")
                summary_sheet.cell(row=row, column=1).font = Font(bold=True, size=12)
                
                row += 1
                summary_sheet.cell(row=row, column=1, value="Total Discrepancies:")
                summary_sheet.cell(row=row, column=2, value=summary['total_discrepancies'])
                if summary['total_discrepancies'] > 0:
                    summary_sheet.cell(row=row, column=2).fill = PatternFill(
                        start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                    )
                else:
                    summary_sheet.cell(row=row, column=2).fill = PatternFill(
                        start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
                    )
                
                row += 1
                summary_sheet.cell(row=row, column=1, value="Position Mismatches:")
                summary_sheet.cell(row=row, column=2, value=summary['mismatch_count'])
                
                row += 1
                summary_sheet.cell(row=row, column=1, value="Missing in Recon:")
                summary_sheet.cell(row=row, column=2, value=summary['missing_in_recon_count'])
                
                row += 1
                summary_sheet.cell(row=row, column=1, value="Missing in Delivery:")
                summary_sheet.cell(row=row, column=2, value=summary['missing_in_delivery_count'])
            
            # Set column widths for summary sheet
            summary_sheet.column_dimensions['A'].width = 35
            summary_sheet.column_dimensions['B'].width = 40
            
            # Save consolidated workbook
            wb_delivery.save(consolidated_file)
            
            logger.info(f"Generated consolidated report: {consolidated_file}")
            return consolidated_file
            
        except Exception as e:
            logger.error(f"Error creating consolidated report: {e}")
            raise
    
    def positions_review_tab(self):
        """Display parsed positions for review"""
        st.markdown('<h2 class="sub-header">Position Summary</h2>', unsafe_allow_html=True)
        
        if not st.session_state.positions:
            st.info("📤 Please upload and process a position file first")
            return
        
        positions = st.session_state.positions
        
        # Summary metrics
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total Positions", len(positions))
        
        with col2:
            unique_underlyings = len(set(p.underlying_ticker for p in positions))
            st.metric("Unique Underlyings", unique_underlyings)
        
        with col3:
            unique_expiries = len(set(p.expiry_date for p in positions))
            st.metric("Unique Expiries", unique_expiries)
        
        with col4:
            futures_count = sum(1 for p in positions if p.is_future)
            options_count = len(positions) - futures_count
            st.metric("Futures/Options", f"{futures_count}/{options_count}")
        
        # Detailed positions table
        st.subheader("📋 Position Details")
        
        # Convert positions to dataframe
        df_data = []
        for p in positions:
            df_data.append({
                'Underlying': p.underlying_ticker,
                'Symbol': p.symbol,
                'Bloomberg Ticker': p.bloomberg_ticker,
                'Expiry': p.expiry_date.strftime('%d/%m/%Y'),
                'Type': p.security_type,
                'Strike': p.strike_price if p.strike_price > 0 else '',
                'Position (Lots)': p.position_lots,
                'Lot Size': p.lot_size
            })
        
        df = pd.DataFrame(df_data)
        
        # Display table
        st.dataframe(
            df,
            use_container_width=True,
            hide_index=True,
            column_config={
                'Strike': st.column_config.NumberColumn(format="%.2f"),
                'Position (Lots)': st.column_config.NumberColumn(format="%.2f"),
            }
        )
        
        # Unmapped symbols warning
        if st.session_state.unmapped_symbols:
            st.warning(f"⚠️ {len(st.session_state.unmapped_symbols)} unmapped symbols found")
            with st.expander("View Unmapped Symbols"):
                unmapped_df = pd.DataFrame(st.session_state.unmapped_symbols)
                st.dataframe(unmapped_df, use_container_width=True, hide_index=True)
    
    def deliverables_preview_tab(self):
        """Preview deliverables calculation"""
        st.markdown('<h2 class="sub-header">Deliverables Analysis</h2>', unsafe_allow_html=True)
        
        if not st.session_state.positions:
            st.info("📤 Please upload and process a position file first")
            return
        
        positions = st.session_state.positions
        prices = st.session_state.prices
        
        # Group by underlying
        grouped = {}
        for p in positions:
            if p.underlying_ticker not in grouped:
                grouped[p.underlying_ticker] = []
            grouped[p.underlying_ticker].append(p)
        
        # Sensitivity analysis
        st.subheader("📈 Sensitivity Analysis")
        sensitivity_pct = st.slider(
            "Price Change %",
            min_value=-20.0,
            max_value=20.0,
            value=0.0,
            step=1.0,
            help="Analyze deliverables at different price levels"
        )
        
        # Calculate deliverables
        deliverables_data = []
        
        for underlying in sorted(grouped.keys()):
            underlying_positions = grouped[underlying]
            spot_price = prices.get(underlying, 0)
            
            if spot_price:
                adjusted_price = spot_price * (1 + sensitivity_pct / 100)
            else:
                adjusted_price = 0
            
            total_deliverable = 0
            
            for pos in underlying_positions:
                if pos.security_type == 'Futures':
                    deliverable = pos.position_lots
                elif pos.security_type == 'Call':
                    if adjusted_price > pos.strike_price:
                        deliverable = pos.position_lots
                    else:
                        deliverable = 0
                elif pos.security_type == 'Put':
                    if adjusted_price < pos.strike_price:
                        deliverable = -pos.position_lots
                    else:
                        deliverable = 0
                else:
                    deliverable = 0
                
                total_deliverable += deliverable
            
            deliverables_data.append({
                'Underlying': underlying,
                'Current Price': spot_price,
                'Adjusted Price': adjusted_price if spot_price else 'N/A',
                'Total Positions': len(underlying_positions),
                'Net Deliverable (Lots)': total_deliverable
            })
        
        # Display table
        deliverables_df = pd.DataFrame(deliverables_data)
        st.dataframe(
            deliverables_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                'Current Price': st.column_config.NumberColumn(format="%.2f"),
                'Adjusted Price': st.column_config.NumberColumn(format="%.2f"),
                'Net Deliverable (Lots)': st.column_config.NumberColumn(format="%.0f"),
            }
        )
    
    def reconciliation_tab(self):
        """Display reconciliation results"""
        st.markdown('<h2 class="sub-header">Position Reconciliation</h2>', unsafe_allow_html=True)
        
        if not st.session_state.report_generated:
            st.info("📤 Please process a position file first")
            return
        
        if not st.session_state.recon_file:
            st.markdown('<div class="recon-box">', unsafe_allow_html=True)
            st.info("📋 Upload a reconciliation file in the sidebar to compare positions")
            st.write("The recon file should have two columns:")
            st.write("- Column A: Symbol (Bloomberg Ticker)")
            st.write("- Column B: Position")
            st.markdown('</div>', unsafe_allow_html=True)
            return
        
        if not st.session_state.recon_results:
            if st.button("🔄 Run Reconciliation", type="primary"):
                with st.spinner("Performing reconciliation..."):
                    self.perform_reconciliation()
        
        if st.session_state.recon_results:
            results = st.session_state.recon_results
            summary = results['summary']
            
            # Display summary metrics
            st.subheader("📊 Reconciliation Summary")
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("Matched Positions", summary['matched_count'])
            
            with col2:
                st.metric("Position Mismatches", summary['mismatch_count'])
            
            with col3:
                st.metric("Missing in Recon", summary['missing_in_recon_count'])
            
            with col4:
                st.metric("Missing in Delivery", summary['missing_in_delivery_count'])
            
            # Show total discrepancies prominently
            if summary['total_discrepancies'] > 0:
                st.error(f"⚠️ Total Discrepancies: {summary['total_discrepancies']}")
            else:
                st.success("✅ All positions match perfectly!")
            
            # Display detailed discrepancies
            if results['position_mismatches']:
                st.subheader("🔍 Position Mismatches")
                mismatch_df = pd.DataFrame(results['position_mismatches'])
                st.dataframe(
                    mismatch_df,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        'Delivery_Position': st.column_config.NumberColumn(format="%.2f"),
                        'Recon_Position': st.column_config.NumberColumn(format="%.2f"),
                        'Difference': st.column_config.NumberColumn(format="%.2f"),
                    }
                )
            
            if results['missing_in_recon']:
                st.subheader("📋 Missing in Recon File")
                missing_recon_df = pd.DataFrame(results['missing_in_recon'])
                st.dataframe(missing_recon_df, use_container_width=True, hide_index=True)
            
            if results['missing_in_delivery']:
                st.subheader("📋 Missing in Delivery Output")
                missing_delivery_df = pd.DataFrame(results['missing_in_delivery'])
                st.dataframe(missing_delivery_df, use_container_width=True, hide_index=True)
    
    def download_reports_tab(self):
        """Download generated reports"""
        st.markdown('<h2 class="sub-header">Download Reports</h2>', unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.subheader("📊 Delivery Report")
            
            if not st.session_state.report_generated or not st.session_state.output_file:
                st.info("📤 Please process a position file first")
            else:
                st.markdown('<div class="success-box">', unsafe_allow_html=True)
                st.success("✅ **Delivery Report Ready!**")
                st.write(f"**Filename:** {st.session_state.output_file}")
                st.markdown('</div>', unsafe_allow_html=True)
                
                try:
                    with open(st.session_state.output_file, 'rb') as f:
                        excel_data = f.read()
                    
                    st.download_button(
                        label="📥 Download Delivery Report",
                        data=excel_data,
                        file_name=st.session_state.output_file,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary"
                    )
                except Exception as e:
                    st.error(f"Error reading report: {str(e)}")
        
        with col2:
            st.subheader("📄 Reconciliation Report")
            
            if not hasattr(st.session_state, 'recon_output_file'):
                st.info("📋 Upload a recon file and run reconciliation first")
            else:
                st.markdown('<div class="success-box">', unsafe_allow_html=True)
                st.success("✅ **Reconciliation Report Ready!**")
                st.write(f"**Filename:** {st.session_state.recon_output_file}")
                st.markdown('</div>', unsafe_allow_html=True)
                
                try:
                    with open(st.session_state.recon_output_file, 'rb') as f:
                        recon_data = f.read()
                    
                    st.download_button(
                        label="📥 Download Reconciliation Report",
                        data=recon_data,
                        file_name=st.session_state.recon_output_file,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary"
                    )
                except Exception as e:
                    st.error(f"Error reading recon report: {str(e)}")
        
        with col3:
            st.subheader("📦 Consolidated Report")
            
            # Check if both reports are available
            if (not st.session_state.report_generated or not st.session_state.output_file or 
                not hasattr(st.session_state, 'recon_output_file')):
                st.info("📋 Generate both reports first")
            else:
                st.markdown('<div class="success-box">', unsafe_allow_html=True)
                st.success("✅ **Both Reports Available!**")
                st.write("Combine delivery and reconciliation reports into a single file")
                st.markdown('</div>', unsafe_allow_html=True)
                
                if st.button("🔄 Generate Consolidated Report", use_container_width=True, type="primary"):
                    try:
                        with st.spinner("Creating consolidated report..."):
                            consolidated_file = self.generate_consolidated_report(
                                st.session_state.output_file,
                                st.session_state.recon_output_file
                            )
                            
                            # Read the consolidated file for download
                            with open(consolidated_file, 'rb') as f:
                                consolidated_data = f.read()
                            
                            st.download_button(
                                label="📥 Download Consolidated Report",
                                data=consolidated_data,
                                file_name=consolidated_file,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                type="secondary"
                            )
                            
                            # Clean up the temporary consolidated file
                            try:
                                os.unlink(consolidated_file)
                            except:
                                pass
                                
                    except Exception as e:
                        st.error(f"Error creating consolidated report: {str(e)}")


def main():
    """Main entry point"""
    app = StreamlitDeliveryApp()
    app.run()


if __name__ == "__main__":
    main()
