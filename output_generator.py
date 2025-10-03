"""
Output Generator Module - UPDATED WITH DD/MM/YYYY DATE FORMAT
All dates in output files are formatted as DD/MM/YYYY except Trade Date in ACM
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional, Set
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# Import email functionality (optional)
try:
    from email_sender import EmailSender
    from email_config import EmailConfig
    EMAIL_AVAILABLE = True
except ImportError:
    EMAIL_AVAILABLE = False
    logger.info("Email functionality not available - install sendgrid package to enable")


class OutputGenerator:
    """Generates and saves all output files with proper date formatting"""

    def __init__(self, output_dir: str = "./output", account_prefix: str = ""):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.missing_mappings = {'positions': [], 'trades': []}
        self.account_prefix = account_prefix  # e.g. "AURIGIN_" or ""
        
    def save_all_outputs(self,
                        parsed_trades_df: pd.DataFrame,
                        starting_positions_df: pd.DataFrame,
                        processed_trades_df: pd.DataFrame,
                        final_positions_df: pd.DataFrame,
                        file_prefix: str = "output",
                        input_parser=None,
                        trade_parser=None,
                        send_email: bool = False,
                        email_recipients: List[str] = None,
                        email_file_filter: Dict[str, bool] = None) -> Dict[str, Path]:
        """
        Save all output files with proper date formatting
        Returns dictionary of file type to file path
        """
        output_files = {}
        
        # Ensure all DataFrames have dates formatted properly
        parsed_trades_df = self._format_dates_in_dataframe(parsed_trades_df)
        starting_positions_df = self._format_dates_in_dataframe(starting_positions_df)
        processed_trades_df = self._format_dates_in_dataframe(processed_trades_df)
        final_positions_df = self._format_dates_in_dataframe(final_positions_df)
        
        # File 1: Parsed Trade File (original trades from parser)
        parsed_trades_file = self.output_dir / f"{self.account_prefix}{file_prefix}_1_parsed_trades_{self.timestamp}.csv"
        parsed_trades_df.to_csv(parsed_trades_file, index=False, date_format='%d/%m/%Y')
        output_files['parsed_trades'] = parsed_trades_file
        logger.info(f"Saved parsed trades to {parsed_trades_file}")
        
        # File 2: Starting Position File
        starting_pos_file = self.output_dir / f"{self.account_prefix}{file_prefix}_2_starting_positions_{self.timestamp}.csv"
        starting_positions_df.to_csv(starting_pos_file, index=False, date_format='%d/%m/%Y')
        output_files['starting_positions'] = starting_pos_file
        logger.info(f"Saved starting positions to {starting_pos_file}")
        
        # File 3: Processed Trade File (main output with strategies)
        processed_trades_file = self.output_dir / f"{self.account_prefix}{file_prefix}_3_processed_trades_{self.timestamp}.csv"
        processed_trades_df.to_csv(processed_trades_file, index=False, date_format='%d/%m/%Y')
        output_files['processed_trades'] = processed_trades_file
        logger.info(f"Saved processed trades to {processed_trades_file}")
        
        # File 4: Final Position File
        final_pos_file = self.output_dir / f"{self.account_prefix}{file_prefix}_4_final_positions_{self.timestamp}.csv"
        final_positions_df.to_csv(final_pos_file, index=False, date_format='%d/%m/%Y')
        output_files['final_positions'] = final_pos_file
        logger.info(f"Saved final positions to {final_pos_file}")
        
        # File 5: Missing Mappings Report
        if input_parser or trade_parser:
            missing_mappings_file = self.create_missing_mappings_report(input_parser, trade_parser)
            if missing_mappings_file:
                output_files['missing_mappings'] = missing_mappings_file
        
        # Create summary report
        summary_file = self._create_summary_report(
            parsed_trades_df, 
            starting_positions_df, 
            processed_trades_df, 
            final_positions_df,
            input_parser,
            trade_parser
        )
        output_files['summary'] = summary_file

        # Send email if requested
        if send_email and email_recipients:
            self._send_completion_email(
                output_files=output_files,
                email_recipients=email_recipients,
                file_filter=email_file_filter,
                stats={
                    'total_trades': len(parsed_trades_df),
                    'starting_positions': len(starting_positions_df),
                    'final_positions': len(final_positions_df)
                }
            )

        return output_files

    def save_positions_by_underlying_excel(self, final_positions_df: pd.DataFrame,
                                          file_prefix: str = "positions_by_underlying",
                                          price_manager=None) -> Optional[Path]:
        """
        Save positions grouped by underlying in hierarchical Excel format
        Similar to excel_writer.py structure
        """
        try:
            # Import position grouper
            from positions_grouper import PositionGrouper

            output_file = self.output_dir / f"{self.account_prefix}{file_prefix}_{self.timestamp}.xlsx"
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Define styles
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            group_font = Font(bold=True, size=10)
            group_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

            # Initialize grouper
            grouper = PositionGrouper()
            grouped_data = grouper.group_positions_from_dataframe(final_positions_df)

            # Add spot prices if available
            if price_manager:
                for underlying in grouped_data:
                    price = price_manager.get_price(underlying)
                    if price:
                        grouped_data[underlying]['spot_price'] = price

            # Create Master sheet with all positions
            ws = wb.create_sheet("Master_All_Positions")

            # Headers
            headers = [
                "Underlying", "Symbol", "Bloomberg Ticker", "Expiry", "Type", "Strike",
                "Position (Lots)", "Lot Size", "Position (Qty)", "Spot Price"
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            current_row = 2
            sorted_underlyings = sorted(grouped_data.keys())

            # Write grouped data
            for underlying in sorted_underlyings:
                data = grouped_data[underlying]
                group_start_row = current_row

                # Write group header
                cell = ws.cell(row=current_row, column=1, value=underlying)
                cell.font = group_font

                for col in range(1, 11):
                    cell = ws.cell(row=current_row, column=col)
                    cell.fill = group_fill
                    cell.border = border

                # Add spot price to group header
                spot_price = data.get('spot_price')
                if spot_price:
                    ws.cell(row=current_row, column=10, value=spot_price).number_format = '#,##0.00'

                # Summary on group header
                ws.cell(row=current_row, column=7, value=data['net_position']).font = group_font

                detail_rows = []
                current_row += 1

                # Write detail rows
                for pos in data['positions']:
                    ws.cell(row=current_row, column=2, value=pos['symbol'])
                    ws.cell(row=current_row, column=3, value=pos['bloomberg_ticker'])
                    ws.cell(row=current_row, column=4,
                           value=pos['expiry'].strftime('%d/%m/%Y') if pos['expiry'] else '')
                    ws.cell(row=current_row, column=5, value=pos['security_type'])

                    if pos['strike']:
                        ws.cell(row=current_row, column=6, value=pos['strike']).number_format = '#,##0.00'

                    ws.cell(row=current_row, column=7, value=pos['position_lots'])
                    ws.cell(row=current_row, column=8, value=pos['lot_size'])
                    ws.cell(row=current_row, column=9, value=pos['position_lots'] * pos['lot_size'])

                    # Apply borders
                    for col in range(1, 11):
                        ws.cell(row=current_row, column=col).border = border

                    detail_rows.append(current_row)
                    current_row += 1

                # Apply grouping (collapsible rows)
                for row_num in detail_rows:
                    ws.row_dimensions[row_num].outline_level = 1
                    ws.row_dimensions[row_num].hidden = False

            # Set column widths
            widths = {
                'A': 25, 'B': 15, 'C': 30, 'D': 12, 'E': 8, 'F': 10,
                'G': 15, 'H': 10, 'I': 15, 'J': 12
            }
            for col, width in widths.items():
                ws.column_dimensions[col].width = width

            # Set outline properties for collapsible groups
            ws.sheet_properties.outlinePr.summaryBelow = False
            ws.sheet_properties.outlinePr.summaryRight = False

            # Create Summary sheet
            ws_summary = wb.create_sheet("Summary", 0)

            # Summary headers
            summary_headers = [
                "Underlying", "Net Position (Lots)", "Futures", "Calls", "Puts",
                "Total Positions", "Unique Expiries", "Spot Price"
            ]

            for col, header in enumerate(summary_headers, 1):
                cell = ws_summary.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Write summary data
            row = 2
            for underlying in sorted_underlyings:
                data = grouped_data[underlying]
                ws_summary.cell(row=row, column=1, value=underlying)
                ws_summary.cell(row=row, column=2, value=data['net_position'])
                ws_summary.cell(row=row, column=3, value=data['total_futures'])
                ws_summary.cell(row=row, column=4, value=data['total_calls'])
                ws_summary.cell(row=row, column=5, value=data['total_puts'])
                ws_summary.cell(row=row, column=6, value=len(data['positions']))
                ws_summary.cell(row=row, column=7, value=len(data['unique_expiries']))

                spot_price = data.get('spot_price')
                if spot_price:
                    ws_summary.cell(row=row, column=8, value=spot_price).number_format = '#,##0.00'

                # Apply borders
                for col in range(1, 9):
                    ws_summary.cell(row=row, column=col).border = border

                row += 1

            # Set summary column widths
            summary_widths = {
                'A': 25, 'B': 18, 'C': 10, 'D': 10, 'E': 10,
                'F': 15, 'G': 15, 'H': 12
            }
            for col, width in summary_widths.items():
                ws_summary.column_dimensions[col].width = width

            # Save workbook
            wb.save(output_file)
            logger.info(f"Saved positions by underlying to {output_file}")

            return output_file

        except Exception as e:
            logger.error(f"Error creating positions by underlying Excel: {e}")
            return None

    def _format_dates_in_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Format date columns in DataFrame to DD/MM/YYYY format
        Handles both datetime objects and strings
        """
        if df.empty:
            return df

        df = df.copy()

        # List of columns that typically contain dates
        date_columns = ['Expiry', 'expiry', 'Expiry_Date', 'expiry_date']

        for col in df.columns:
            if col in date_columns or 'expiry' in col.lower():
                if col in df.columns:
                    # Convert to datetime first if not already
                    try:
                        df[col] = pd.to_datetime(df[col])
                        # Format as DD/MM/YYYY date string
                        df[col] = df[col].dt.strftime('%d/%m/%Y')
                    except:
                        # If conversion fails, check if it's already a string in correct format
                        pass

        return df
    
    def create_missing_mappings_report(self, input_parser=None, trade_parser=None) -> Optional[Path]:
        """
        Create a report of all unmapped symbols from both parsers
        Returns the path to the CSV file
        """
        all_missing = []
        
        # Collect from input parser (positions)
        if input_parser and hasattr(input_parser, 'unmapped_symbols'):
            for item in input_parser.unmapped_symbols:
                # Format expiry as DD/MM/YYYY if it exists
                expiry = item.get('expiry', '')
                if expiry and hasattr(expiry, 'strftime'):
                    expiry = expiry.strftime('%d/%m/%Y')
                
                all_missing.append({
                    'Source': 'Position File',
                    'Symbol': item.get('symbol', ''),
                    'Expiry': expiry,
                    'Quantity': item.get('position_lots', 0),
                    'Suggested_Ticker': self._suggest_ticker(item.get('symbol', '')),
                    'Underlying': '',
                    'Exchange': '',
                    'Lot_Size': ''
                })
        
        # Collect from trade parser
        if trade_parser and hasattr(trade_parser, 'unmapped_symbols'):
            for item in trade_parser.unmapped_symbols:
                # Format expiry as DD/MM/YYYY if it exists
                expiry = item.get('expiry', '')
                if expiry and hasattr(expiry, 'strftime'):
                    expiry = expiry.strftime('%d/%m/%Y')
                    
                all_missing.append({
                    'Source': 'Trade File',
                    'Symbol': item.get('symbol', ''),
                    'Expiry': expiry,
                    'Quantity': item.get('position_lots', 0),
                    'Suggested_Ticker': self._suggest_ticker(item.get('symbol', '')),
                    'Underlying': '',
                    'Exchange': '',
                    'Lot_Size': ''
                })
        
        if not all_missing:
            logger.info("No missing mappings found")
            return None
        
        # Create DataFrame and remove duplicates
        df = pd.DataFrame(all_missing)
        
        # Group by symbol to consolidate
        unique_symbols = df.groupby('Symbol').agg({
            'Source': lambda x: ', '.join(sorted(set(x))),
            'Expiry': 'first',
            'Quantity': 'sum',
            'Suggested_Ticker': 'first',
            'Underlying': 'first',
            'Exchange': 'first',
            'Lot_Size': 'first'
        }).reset_index()
        
        # Sort by symbol
        unique_symbols = unique_symbols.sort_values('Symbol')
        
        # Save to CSV
        missing_file = self.output_dir / f"{self.account_prefix}MISSING_MAPPINGS_{self.timestamp}.csv"
        unique_symbols.to_csv(missing_file, index=False, date_format='%d/%m/%Y')

        # Also create a template for easy addition to mapping file
        template_file = self.output_dir / f"{self.account_prefix}MAPPING_TEMPLATE_{self.timestamp}.csv"
        template_df = unique_symbols[['Symbol', 'Suggested_Ticker', 'Underlying', 'Exchange', 'Lot_Size']]
        template_df.columns = ['Symbol', 'Ticker', 'Underlying', 'Exchange', 'Lot_Size']
        template_df.to_csv(template_file, index=False)
        
        logger.info(f"Created missing mappings report with {len(unique_symbols)} unmapped symbols")
        logger.info(f"Missing mappings report: {missing_file}")
        logger.info(f"Mapping template file: {template_file}")
        
        return missing_file
    
    def _suggest_ticker(self, symbol: str) -> str:
        """
        Suggest a ticker based on common patterns
        """
        symbol_upper = symbol.upper()
        
        # Remove common suffixes
        cleaned = symbol_upper
        for suffix in ['EQ', 'FUT', 'OPT', 'CE', 'PE', '-EQ', '-FUT']:
            if cleaned.endswith(suffix):
                cleaned = cleaned[:-len(suffix)]
                break
        
        # Common index mappings
        index_map = {
            'NIFTY': 'NZ',
            'BANKNIFTY': 'AF1',
            'FINNIFTY': 'FINNIFTY',
            'MIDCPNIFTY': 'RNS'
        }
        
        for key, value in index_map.items():
            if key in symbol_upper:
                return value
        
        # For others, return cleaned version
        return cleaned.strip('-').strip()
    
    def _create_summary_report(self,
                              parsed_trades_df: pd.DataFrame,
                              starting_positions_df: pd.DataFrame,
                              processed_trades_df: pd.DataFrame,
                              final_positions_df: pd.DataFrame,
                              input_parser=None,
                              trade_parser=None) -> Path:
        """Create a summary report of the processing including missing mappings"""
        summary_file = self.output_dir / f"{self.account_prefix}summary_report_{self.timestamp}.txt"

        with open(summary_file, 'w', encoding='utf-8') as f:
            f.write("=" * 60 + "\n")
            f.write("TRADE PROCESSING SUMMARY REPORT\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 60 + "\n\n")
            
            # Missing mappings section
            missing_count = 0
            if input_parser and hasattr(input_parser, 'unmapped_symbols'):
                missing_count += len(input_parser.unmapped_symbols)
            if trade_parser and hasattr(trade_parser, 'unmapped_symbols'):
                missing_count += len(trade_parser.unmapped_symbols)
            
            if missing_count > 0:
                f.write("WARNING: MISSING MAPPINGS:\n")
                f.write("-" * 30 + "\n")
                
                if input_parser and hasattr(input_parser, 'unmapped_symbols') and input_parser.unmapped_symbols:
                    f.write(f"Position file: {len(input_parser.unmapped_symbols)} unmapped symbols\n")
                    unique_pos_symbols = set(item['symbol'] for item in input_parser.unmapped_symbols)
                    f.write(f"  Symbols: {', '.join(sorted(unique_pos_symbols)[:10])}")
                    if len(unique_pos_symbols) > 10:
                        f.write(f" ... and {len(unique_pos_symbols) - 10} more")
                    f.write("\n")
                
                if trade_parser and hasattr(trade_parser, 'unmapped_symbols') and trade_parser.unmapped_symbols:
                    f.write(f"Trade file: {len(trade_parser.unmapped_symbols)} unmapped symbols\n")
                    unique_trade_symbols = set(item['symbol'] for item in trade_parser.unmapped_symbols)
                    f.write(f"  Symbols: {', '.join(sorted(unique_trade_symbols)[:10])}")
                    if len(unique_trade_symbols) > 10:
                        f.write(f" ... and {len(unique_trade_symbols) - 10} more")
                    f.write("\n")
                
                f.write("\nNote: Check MISSING_MAPPINGS_*.csv for complete list\n")
                f.write("Note: Use MAPPING_TEMPLATE_*.csv to add to your mapping file\n\n")
            
            # Starting positions summary
            f.write("STARTING POSITIONS:\n")
            f.write("-" * 30 + "\n")
            f.write(f"Total positions: {len(starting_positions_df)}\n")
            if len(starting_positions_df) > 0:
                f.write(f"Long positions: {len(starting_positions_df[starting_positions_df['QTY'] > 0])}\n")
                f.write(f"Short positions: {len(starting_positions_df[starting_positions_df['QTY'] < 0])}\n")
            f.write("\n")
            
            # Trades summary
            f.write("TRADES PROCESSED:\n")
            f.write("-" * 30 + "\n")
            f.write(f"Total trades: {len(parsed_trades_df)}\n")
            f.write(f"Trades after processing: {len(processed_trades_df)}\n")
            
            # Split trades
            if 'Split?' in processed_trades_df.columns:
                split_trades = processed_trades_df[processed_trades_df['Split?'] == 'Yes']
                f.write(f"Split trades: {len(split_trades)} (from {len(split_trades)//2} original trades)\n")
            
            # Opposite trades
            if 'Opposite?' in processed_trades_df.columns:
                opposite_trades = processed_trades_df[processed_trades_df['Opposite?'] == 'Yes']
                f.write(f"Trades with opposite strategy: {len(opposite_trades)}\n")
            
            # Strategy breakdown
            if 'Strategy' in processed_trades_df.columns:
                f.write("\nStrategy Breakdown:\n")
                strategy_counts = processed_trades_df['Strategy'].value_counts()
                for strategy, count in strategy_counts.items():
                    f.write(f"  {strategy}: {count} trades\n")
            
            f.write("\n")
            
            # Final positions summary
            f.write("FINAL POSITIONS:\n")
            f.write("-" * 30 + "\n")
            f.write(f"Total positions: {len(final_positions_df)}\n")
            if len(final_positions_df) > 0:
                f.write(f"Long positions: {len(final_positions_df[final_positions_df['QTY'] > 0])}\n")
                f.write(f"Short positions: {len(final_positions_df[final_positions_df['QTY'] < 0])}\n")
                
                # Position changes
                f.write("\nPosition Changes:\n")
                initial_tickers = set(starting_positions_df['Ticker'].unique()) if len(starting_positions_df) > 0 else set()
                final_tickers = set(final_positions_df['Ticker'].unique()) if len(final_positions_df) > 0 else set()
                
                new_positions = final_tickers - initial_tickers
                closed_positions = initial_tickers - final_tickers
                
                if new_positions:
                    f.write(f"  New positions opened: {len(new_positions)}\n")
                if closed_positions:
                    f.write(f"  Positions closed: {len(closed_positions)}\n")
            
            # Note about date formatting
            f.write("\n")
            f.write("DATE FORMAT NOTES:\n")
            f.write("-" * 30 + "\n")
            f.write("All dates in Stage 1 outputs: DD/MM/YYYY (date format)\n")
            f.write("ACM Trade Date (Stage 2): MM/DD/YYYY HH:MM:SS (datetime format)\n")
            f.write("ACM Settle Date (Stage 2): MM/DD/YYYY (date format)\n")
            
            f.write("\n" + "=" * 60 + "\n")
            f.write("END OF REPORT\n")
            f.write("=" * 60 + "\n")
        
        logger.info(f"Created summary report at {summary_file}")
        return summary_file
    
    def create_trade_dataframe_from_positions(self, positions: List) -> pd.DataFrame:
        """Convert Position objects to DataFrame for parsed trades output with DD/MM/YYYY date format"""
        trades_data = []

        for pos in positions:
            # Format expiry as DD/MM/YYYY
            expiry_str = pos.expiry_date.strftime('%d/%m/%Y')
            
            trade_dict = {
                'Symbol': pos.symbol,
                'Bloomberg_Ticker': pos.bloomberg_ticker,
                'Expiry': expiry_str,  # Simple date format
                'Strike': pos.strike_price,
                'Security_Type': pos.security_type,
                'Lots': pos.position_lots,
                'Lot_Size': pos.lot_size,
                'Quantity': pos.position_lots * pos.lot_size
            }
            trades_data.append(trade_dict)
        
        return pd.DataFrame(trades_data)

    def _send_completion_email(self, output_files: Dict[str, Path], email_recipients: List[str],
                               file_filter: Dict[str, bool], stats: Dict):
        """Send email notification when processing is complete"""
        if not email_recipients:
            logger.info("No email recipients specified - skipping email")
            return

        if not EMAIL_AVAILABLE:
            logger.warning("Email not sent - sendgrid package not installed")
            logger.warning("Run: pip install sendgrid python-dotenv")
            return

        try:
            logger.info(f"Attempting to send email to: {', '.join(email_recipients)}")
            email_sender = EmailSender()

            if not email_sender.is_enabled():
                logger.warning("Email not sent - SendGrid not configured")
                logger.warning("Check .env file for SENDGRID_API_KEY and SENDGRID_FROM_EMAIL")
                return

            # Send Stage 1 completion email
            logger.info("Sending Stage 1 completion email...")
            success = email_sender.send_stage1_complete(
                to_emails=email_recipients,
                account_prefix=self.account_prefix,
                timestamp=self.timestamp,
                output_files=output_files,
                stats=stats,
                file_filter=file_filter
            )

            if success:
                logger.info(f"✅ Email sent successfully to {', '.join(email_recipients)}")
            else:
                logger.error("❌ Failed to send completion email")

        except Exception as e:
            logger.error(f"❌ Error sending completion email: {e}")
            import traceback
            logger.error(traceback.format_exc())
