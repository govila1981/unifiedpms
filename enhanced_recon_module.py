"""
Enhanced Position Reconciliation Module
Compares positions from trade processing output with external PMS files
Works with both pre-trade and post-trade positions
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from typing import List, Dict, Tuple, Optional
import logging

logger = logging.getLogger(__name__)


class EnhancedReconciliation:
    """Enhanced reconciliation for trade processing system"""
    
    def __init__(self):
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.mismatch_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        self.missing_fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
        self.extra_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        self.match_fill = PatternFill(start_color="E5FFE5", end_color="E5FFE5", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.header_alignment = Alignment(horizontal="center", vertical="center")
    
    def read_pms_file(self, file_path: str) -> pd.DataFrame:
        """
        Read PMS reconciliation file (Excel or CSV)
        Expects columns: Symbol/Ticker and Position/Quantity
        
        Returns:
            DataFrame with standardized columns [Symbol, Position]
        """
        try:
            # Determine file type
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                # Excel file - read first sheet
                df = pd.read_excel(file_path)
            
            # Standardize column names
            # Look for symbol/ticker column
            symbol_col = None
            position_col = None
            
            for col in df.columns:
                col_lower = col.lower()
                if 'symbol' in col_lower or 'ticker' in col_lower:
                    symbol_col = col
                elif 'position' in col_lower or 'qty' in col_lower or 'quantity' in col_lower:
                    position_col = col
            
            if symbol_col is None or position_col is None:
                # Use first two columns as fallback
                symbol_col = df.columns[0]
                position_col = df.columns[1]
                logger.warning(f"Could not identify columns, using {symbol_col} and {position_col}")
            
            # Create standardized dataframe
            pms_df = pd.DataFrame({
                'Symbol': df[symbol_col].astype(str).str.strip(),
                'Position': pd.to_numeric(df[position_col], errors='coerce')
            })
            
            # Remove NaN rows
            pms_df = pms_df.dropna()
            
            logger.info(f"Read {len(pms_df)} positions from PMS file")
            return pms_df
            
        except Exception as e:
            logger.error(f"Error reading PMS file: {e}")
            raise
    
    def reconcile_positions(self, system_df: pd.DataFrame, pms_df: pd.DataFrame, 
                           position_type: str = "Current") -> Dict:
        """
        Compare system positions with PMS positions
        
        Args:
            system_df: DataFrame from trade processing system [Ticker, Lots, Symbol, etc.]
            pms_df: DataFrame from PMS [Symbol, Position]
            position_type: Type of positions (Pre-Trade, Post-Trade, Current)
            
        Returns:
            Dictionary with reconciliation results
        """
        # Prepare system dataframe - use Ticker as the key
        if 'Ticker' in system_df.columns:
            system_recon = pd.DataFrame({
                'Symbol': system_df['Ticker'].astype(str).str.strip(),
                'Position': pd.to_numeric(system_df.get('Lots', system_df.get('QTY', 0)), errors='coerce')
            })
        else:
            # Fallback if no Ticker column
            system_recon = pd.DataFrame({
                'Symbol': system_df.iloc[:, 0].astype(str).str.strip(),
                'Position': pd.to_numeric(system_df.iloc[:, 1], errors='coerce')
            })
        
        # Remove zero positions
        system_recon = system_recon[system_recon['Position'] != 0]
        
        # Merge to find matches and differences
        merged = pd.merge(
            system_recon,
            pms_df,
            on='Symbol',
            how='outer',
            suffixes=('_System', '_PMS'),
            indicator=True
        )
        
        results = {
            'position_type': position_type,
            'matched_positions': [],
            'position_mismatches': [],
            'missing_in_pms': [],
            'missing_in_system': [],
            'summary': {}
        }
        
        for _, row in merged.iterrows():
            symbol = row['Symbol']
            pos_system = row.get('Position_System', 0)
            pos_pms = row.get('Position_PMS', 0)
            
            # Handle NaN values
            pos_system = 0 if pd.isna(pos_system) else float(pos_system)
            pos_pms = 0 if pd.isna(pos_pms) else float(pos_pms)
            
            if row['_merge'] == 'both':
                # Symbol exists in both
                if abs(pos_system - pos_pms) < 0.0001:  # Consider floating point precision
                    results['matched_positions'].append({
                        'Symbol': symbol,
                        'Position': pos_system
                    })
                else:
                    results['position_mismatches'].append({
                        'Symbol': symbol,
                        'System_Position': pos_system,
                        'PMS_Position': pos_pms,
                        'Difference': pos_system - pos_pms
                    })
            elif row['_merge'] == 'left_only':
                # In system but not in PMS
                results['missing_in_pms'].append({
                    'Symbol': symbol,
                    'System_Position': pos_system
                })
            else:  # right_only
                # In PMS but not in system
                results['missing_in_system'].append({
                    'Symbol': symbol,
                    'PMS_Position': pos_pms
                })
        
        # Calculate summary
        results['summary'] = {
            'position_type': position_type,
            'total_system_positions': len(system_recon),
            'total_pms_positions': len(pms_df),
            'matched_count': len(results['matched_positions']),
            'mismatch_count': len(results['position_mismatches']),
            'missing_in_pms_count': len(results['missing_in_pms']),
            'missing_in_system_count': len(results['missing_in_system']),
            'total_discrepancies': (
                len(results['position_mismatches']) +
                len(results['missing_in_pms']) +
                len(results['missing_in_system'])
            )
        }
        
        return results
    
    def create_comprehensive_recon_report(self, 
                                         starting_positions_df: pd.DataFrame,
                                         final_positions_df: pd.DataFrame,
                                         pms_df: pd.DataFrame,
                                         output_file: str) -> str:
        """
        Create comprehensive reconciliation report comparing both pre and post trade positions
        
        Returns:
            Path to generated Excel file
        """
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Reconcile both position sets
        pre_trade_recon = self.reconcile_positions(starting_positions_df, pms_df, "Pre-Trade")
        post_trade_recon = self.reconcile_positions(final_positions_df, pms_df, "Post-Trade")
        
        # 1. Executive Summary
        ws_summary = wb.create_sheet("Executive_Summary")
        self._write_executive_summary(ws_summary, pre_trade_recon, post_trade_recon)
        
        # 2. Pre-Trade Reconciliation
        self._add_recon_sheets(wb, pre_trade_recon, "PreTrade")
        
        # 3. Post-Trade Reconciliation
        self._add_recon_sheets(wb, post_trade_recon, "PostTrade")
        
        # 4. Impact Analysis
        ws_impact = wb.create_sheet("Trade_Impact_Analysis")
        self._write_impact_analysis(ws_impact, pre_trade_recon, post_trade_recon)
        
        # Save workbook
        wb.save(output_file)
        logger.info(f"Reconciliation report saved: {output_file}")
        return output_file
    
    def _write_executive_summary(self, ws, pre_recon: Dict, post_recon: Dict):
        """Write executive summary of reconciliation"""
        ws.cell(row=1, column=1, value="POSITION RECONCILIATION EXECUTIVE SUMMARY").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Generated: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M:%S')}")
        
        row = 4
        
        # Pre-Trade Summary
        ws.cell(row=row, column=1, value="PRE-TRADE RECONCILIATION").font = Font(bold=True, size=12)
        row += 1
        
        pre_summary = pre_recon['summary']
        summary_items = [
            ("System Positions", pre_summary['total_system_positions']),
            ("PMS Positions", pre_summary['total_pms_positions']),
            ("Matched", pre_summary['matched_count']),
            ("Mismatches", pre_summary['mismatch_count']),
            ("Missing in PMS", pre_summary['missing_in_pms_count']),
            ("Missing in System", pre_summary['missing_in_system_count']),
            ("Total Discrepancies", pre_summary['total_discrepancies'])
        ]
        
        for label, value in summary_items:
            ws.cell(row=row, column=1, value=label)
            cell = ws.cell(row=row, column=2, value=value)
            
            if label == "Total Discrepancies":
                if value > 0:
                    cell.fill = self.mismatch_fill
                else:
                    cell.fill = self.match_fill
            
            row += 1
        
        # Post-Trade Summary
        row += 1
        ws.cell(row=row, column=1, value="POST-TRADE RECONCILIATION").font = Font(bold=True, size=12)
        row += 1
        
        post_summary = post_recon['summary']
        summary_items = [
            ("System Positions", post_summary['total_system_positions']),
            ("PMS Positions", post_summary['total_pms_positions']),
            ("Matched", post_summary['matched_count']),
            ("Mismatches", post_summary['mismatch_count']),
            ("Missing in PMS", post_summary['missing_in_pms_count']),
            ("Missing in System", post_summary['missing_in_system_count']),
            ("Total Discrepancies", post_summary['total_discrepancies'])
        ]
        
        for label, value in summary_items:
            ws.cell(row=row, column=1, value=label)
            cell = ws.cell(row=row, column=2, value=value)
            
            if label == "Total Discrepancies":
                if value > 0:
                    cell.fill = self.mismatch_fill
                else:
                    cell.fill = self.match_fill
            
            row += 1
        
        # Change Analysis
        row += 1
        ws.cell(row=row, column=1, value="RECONCILIATION CHANGES").font = Font(bold=True, size=12)
        row += 1
        
        disc_change = post_summary['total_discrepancies'] - pre_summary['total_discrepancies']
        ws.cell(row=row, column=1, value="Discrepancy Change")
        cell = ws.cell(row=row, column=2, value=disc_change)
        
        if disc_change > 0:
            cell.fill = self.mismatch_fill  # More discrepancies
        elif disc_change < 0:
            cell.fill = self.match_fill  # Fewer discrepancies
        
        # Apply borders
        for r in range(4, row + 1):
            for c in range(1, 3):
                ws.cell(row=r, column=c).border = self.border
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def _add_recon_sheets(self, wb, recon_results: Dict, prefix: str):
        """Add reconciliation sheets for a position set"""
        
        # Position Mismatches
        if recon_results['position_mismatches']:
            ws = wb.create_sheet(f"{prefix}_Mismatches")
            self._write_mismatches(ws, recon_results['position_mismatches'], prefix)
        
        # Missing in PMS
        if recon_results['missing_in_pms']:
            ws = wb.create_sheet(f"{prefix}_Missing_PMS")
            self._write_missing(ws, recon_results['missing_in_pms'], 'PMS', 'System')
        
        # Missing in System
        if recon_results['missing_in_system']:
            ws = wb.create_sheet(f"{prefix}_Missing_System")
            self._write_missing(ws, recon_results['missing_in_system'], 'System', 'PMS')
        
        # Matched Positions (optional)
        if recon_results['matched_positions']:
            ws = wb.create_sheet(f"{prefix}_Matched")
            self._write_matched(ws, recon_results['matched_positions'])
    
    def _write_mismatches(self, ws, mismatches: List[Dict], position_type: str):
        """Write position mismatches sheet"""
        ws.cell(row=1, column=1, value=f"{position_type.upper()} POSITION MISMATCHES").font = Font(bold=True, size=12)
        
        headers = ["Symbol", "System Position", "PMS Position", "Difference", "Abs Difference"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        row = 4
        for item in sorted(mismatches, key=lambda x: abs(x['Difference']), reverse=True):
            ws.cell(row=row, column=1, value=item['Symbol'])
            ws.cell(row=row, column=2, value=item['System_Position'])
            ws.cell(row=row, column=3, value=item['PMS_Position'])
            ws.cell(row=row, column=4, value=item['Difference'])
            ws.cell(row=row, column=5, value=abs(item['Difference']))
            
            # Highlight the row
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = self.mismatch_fill
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        for col in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 18
    
    def _write_missing(self, ws, missing_items: List[Dict], missing_in: str, has_in: str):
        """Write missing positions sheet"""
        ws.cell(row=1, column=1, value=f"POSITIONS MISSING IN {missing_in.upper()}").font = Font(bold=True, size=12)
        
        headers = ["Symbol", f"{has_in} Position"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        row = 4
        for item in sorted(missing_items, key=lambda x: x['Symbol']):
            ws.cell(row=row, column=1, value=item['Symbol'])
            
            # Get position value from correct key
            if f'{has_in}_Position' in item:
                pos_value = item[f'{has_in}_Position']
            else:
                # Fallback to any position key
                for key in item.keys():
                    if 'Position' in key:
                        pos_value = item[key]
                        break
                else:
                    pos_value = 0
            
            ws.cell(row=row, column=2, value=pos_value)
            
            # Highlight the row
            fill_color = self.missing_fill if missing_in == 'PMS' else self.extra_fill
            for col in range(1, 3):
                ws.cell(row=row, column=col).fill = fill_color
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
    
    def _write_matched(self, ws, matched_items: List[Dict]):
        """Write matched positions sheet"""
        ws.cell(row=1, column=1, value="MATCHED POSITIONS").font = Font(bold=True, size=12)
        
        headers = ["Symbol", "Position"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        row = 4
        for item in sorted(matched_items, key=lambda x: x['Symbol']):
            ws.cell(row=row, column=1, value=item['Symbol'])
            ws.cell(row=row, column=2, value=item['Position'])
            
            for col in range(1, 3):
                ws.cell(row=row, column=col).fill = self.match_fill
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
    
    def _write_impact_analysis(self, ws, pre_recon: Dict, post_recon: Dict):
        """Write trade impact analysis showing how trades affected discrepancies"""
        ws.cell(row=1, column=1, value="TRADE IMPACT ON RECONCILIATION").font = Font(bold=True, size=12)
        
        # Analyze what changed
        pre_mismatches = {m['Symbol']: m for m in pre_recon['position_mismatches']}
        post_mismatches = {m['Symbol']: m for m in post_recon['position_mismatches']}
        
        # Find improvements and deteriorations
        improvements = []
        deteriorations = []
        unchanged = []
        
        all_symbols = set(list(pre_mismatches.keys()) + list(post_mismatches.keys()))
        
        for symbol in all_symbols:
            pre_diff = abs(pre_mismatches[symbol]['Difference']) if symbol in pre_mismatches else 0
            post_diff = abs(post_mismatches[symbol]['Difference']) if symbol in post_mismatches else 0
            
            change = post_diff - pre_diff
            
            if abs(change) < 0.0001:
                unchanged.append(symbol)
            elif change < 0:
                improvements.append({
                    'Symbol': symbol,
                    'Pre_Diff': pre_diff,
                    'Post_Diff': post_diff,
                    'Improvement': abs(change)
                })
            else:
                deteriorations.append({
                    'Symbol': symbol,
                    'Pre_Diff': pre_diff,
                    'Post_Diff': post_diff,
                    'Deterioration': change
                })
        
        row = 3
        
        # Write improvements
        if improvements:
            ws.cell(row=row, column=1, value="IMPROVEMENTS").font = Font(bold=True, color="008000")
            row += 1
            
            headers = ["Symbol", "Pre-Trade Diff", "Post-Trade Diff", "Improvement"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
            row += 1
            
            for item in improvements:
                ws.cell(row=row, column=1, value=item['Symbol'])
                ws.cell(row=row, column=2, value=item['Pre_Diff'])
                ws.cell(row=row, column=3, value=item['Post_Diff'])
                ws.cell(row=row, column=4, value=item['Improvement'])
                
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = self.match_fill
                    ws.cell(row=row, column=col).border = self.border
                row += 1
        
        row += 1
        
        # Write deteriorations
        if deteriorations:
            ws.cell(row=row, column=1, value="DETERIORATIONS").font = Font(bold=True, color="FF0000")
            row += 1
            
            headers = ["Symbol", "Pre-Trade Diff", "Post-Trade Diff", "Deterioration"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
            row += 1
            
            for item in deteriorations:
                ws.cell(row=row, column=1, value=item['Symbol'])
                ws.cell(row=row, column=2, value=item['Pre_Diff'])
                ws.cell(row=row, column=3, value=item['Post_Diff'])
                ws.cell(row=row, column=4, value=item['Deterioration'])
                
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = self.mismatch_fill
                    ws.cell(row=row, column=col).border = self.border
                row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        for col in ['B', 'C', 'D']:
            ws.column_dimensions[col].width = 18
