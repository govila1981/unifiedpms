"""
Excel Writer Module - Complete Version
Handles all Excel output formatting and sheet creation
All calculations happen in Excel via formulas
"""

import logging
from datetime import datetime
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from input_parser import Position

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Write Excel file with grouping and formatting"""
    
    def __init__(self, output_file: str, usdinr_rate: float = 88.0):
        self.output_file = output_file
        self.usdinr_rate = usdinr_rate
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        
        # Define styles
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.group_font = Font(bold=True, size=10)
        self.group_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        
        # Number formats
        self.price_format = '#,##0.00'
        self.deliv_format = '#,##0'
        self.iv_format = '#,##0'
        self.percent_format = '0.00%'
    
    def create_report(self, positions: List[Position], prices: Dict[str, float], 
                     unmapped_symbols: List[Dict]):
        """Create complete Excel report with all sheets"""
        
        # Write deliverable sheets
        self.write_master_sheet(positions, prices)
        
        # Write expiry-wise deliverable sheets
        expiries = list(set(p.expiry_date for p in positions))
        for expiry in sorted(expiries):
            self.write_expiry_sheet(expiry, positions, prices)
        
        # Write IV sheets
        self.write_iv_master_sheet(positions, prices)
        
        # Write expiry-wise IV sheets
        for expiry in sorted(expiries):
            self.write_iv_expiry_sheet(expiry, positions, prices)
        
        # Write all positions sheet
        self.write_all_positions_sheet(positions)
        
        # Write unmapped symbols if any
        if unmapped_symbols:
            self.write_unmapped_sheet(unmapped_symbols)
        
        # Save file
        self.save()
    
    def write_master_sheet(self, positions: List[Position], prices: Dict[str, float]):
        """Write master sheet with all expiries INCLUDING SENSITIVITY COLUMNS"""
        ws = self.wb.create_sheet("Master_All_Expiries")
        
        # Write headers
        headers = [
            "Underlying", "Symbol", "Expiry", "Position", "Type", "Strike",
            "System Deliverable", "Override Deliverable", "System Price",
            "Override Price", "BBG Price", "BBG Deliverable",
            "", "System Price -%", "System Deliv -%", "System Price +%", "System Deliv +%",
            "BBG Price -%", "BBG Deliv -%", "BBG Price +%", "BBG Deliv +%"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            if col == 13:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            else:
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Group positions by underlying
        grouped = self._group_positions(positions)
        sorted_underlyings = sorted(grouped.keys())
        
        current_row = 2
        
        for underlying in sorted_underlyings:
            underlying_positions = grouped[underlying]
            group_start_row = current_row
            
            # Write group header
            cell = ws.cell(row=current_row, column=1, value=underlying)
            cell.font = self.group_font
            
            for col in range(1, 22):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = self.group_fill
                cell.border = self.border
            
            spot_price = prices.get(underlying)
            
            # Prices on group header
            price_cell = ws.cell(row=current_row, column=9, value=spot_price if spot_price else "")
            if spot_price:
                price_cell.number_format = self.price_format
            
            override_cell = ws.cell(row=current_row, column=10, value="")
            override_cell.number_format = self.price_format
            
            bbg_formula = f'=BDP(A{current_row},"PX_LAST")'
            bbg_cell = ws.cell(row=current_row, column=11, value=bbg_formula)
            bbg_cell.number_format = self.price_format
            
            # Sensitivity price formulas
            ws.cell(row=current_row, column=14, 
                value=f"=IF($M$1<>\"\",I{current_row}*(1-$M$1/100),\"\")").number_format = self.price_format
            ws.cell(row=current_row, column=16,
                value=f"=IF($M$1<>\"\",I{current_row}*(1+$M$1/100),\"\")").number_format = self.price_format
            ws.cell(row=current_row, column=18,
                value=f"=IF($M$1<>\"\",K{current_row}*(1-$M$1/100),\"\")").number_format = self.price_format
            ws.cell(row=current_row, column=20,
                value=f"=IF($M$1<>\"\",K{current_row}*(1+$M$1/100),\"\")").number_format = self.price_format
            
            detail_rows = []
            current_row += 1

            # Write detail rows
            for pos in underlying_positions:
                ws.cell(row=current_row, column=2, value=pos.bloomberg_ticker)
                ws.cell(row=current_row, column=3, value=pos.expiry_date.strftime('%d/%m/%Y'))
                ws.cell(row=current_row, column=4, value=pos.position_lots)
                ws.cell(row=current_row, column=5, value=pos.security_type)
                
                strike_cell = ws.cell(row=current_row, column=6, 
                    value=pos.strike_price if pos.strike_price > 0 else "")
                if pos.strike_price > 0:
                    strike_cell.number_format = self.price_format
                
                # Deliverable formulas
                for col_idx, price_col in [(7, "I"), (8, "J"), (12, "K")]:
                    formula = self._create_deliverable_formula(
                        current_row, group_start_row, pos.security_type, 
                        pos.strike_price, pos.position_lots, price_col
                    )
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                    cell.number_format = self.deliv_format
                
                # Sensitivity deliverable formulas
                for col_idx, price_col in [(15, "N"), (17, "P"), (19, "R"), (21, "T")]:
                    base_formula = self._create_deliverable_formula(
                        current_row, group_start_row, pos.security_type,
                        pos.strike_price, pos.position_lots, price_col
                    )
                    cell = ws.cell(row=current_row, column=col_idx, 
                        value=f"=IF($M$1<>\"\",{base_formula[1:]},\"\")")
                    cell.number_format = self.deliv_format
                
                for col in range(1, 22):
                    ws.cell(row=current_row, column=col).border = self.border
                
                detail_rows.append(current_row)
                current_row += 1
            
            # Group header totals
            if detail_rows:
                for col_idx in [7, 8, 12]:
                    col_letter = chr(64 + col_idx)
                    cell = ws.cell(row=group_start_row, column=col_idx,
                        value=f"=SUM({col_letter}{detail_rows[0]}:{col_letter}{detail_rows[-1]})")
                    cell.number_format = self.deliv_format
                
                for col_idx, col_letter in [(15, "O"), (17, "Q"), (19, "S"), (21, "U")]:
                    cell = ws.cell(row=group_start_row, column=col_idx,
                        value=f"=IF($M$1<>\"\",SUM({col_letter}{detail_rows[0]}:{col_letter}{detail_rows[-1]}),\"\")")
                    cell.number_format = self.deliv_format
            
            # Apply grouping
            for row_num in detail_rows:
                ws.row_dimensions[row_num].outline_level = 1
                ws.row_dimensions[row_num].hidden = False
        
        self._set_master_column_widths(ws)
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.summaryRight = False
    
    def write_expiry_sheet(self, expiry_date: datetime, positions: List[Position], prices: Dict[str, float]):
        """Write sheet for specific expiry"""
        sheet_name = f"Expiry_{expiry_date.strftime('%d_%m_%Y')}"
        ws = self.wb.create_sheet(sheet_name)
        
        expiry_positions = [p for p in positions if p.expiry_date.date() == expiry_date.date()]
        
        if not expiry_positions:
            return
        
        # Write headers (extended with sensitivity columns)
        headers = [
            "Underlying", "Symbol", "Expiry", "Position", "Type", "Strike",
            "System Deliverable", "Override Deliverable", "System Price",
            "Override Price", "BBG Price", "BBG Deliverable",
            "", "System Price -%", "System Deliv -%", "System Price +%", "System Deliv +%",
            "BBG Price -%", "BBG Deliv -%", "BBG Price +%", "BBG Deliv +%"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            if col == 13:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            else:
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Group positions by underlying
        grouped = self._group_positions(expiry_positions)
        sorted_underlyings = sorted(grouped.keys())
        
        current_row = 2
        
        for underlying in sorted_underlyings:
            underlying_positions = grouped[underlying]
            group_start_row = current_row
            
            # Write group header
            cell = ws.cell(row=current_row, column=1, value=underlying)
            cell.font = self.group_font
            
            for col in range(1, 22):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = self.group_fill
                cell.border = self.border
            
            spot_price = prices.get(underlying)
            
            # Prices on group header
            price_cell = ws.cell(row=current_row, column=9, value=spot_price if spot_price else "")
            if spot_price:
                price_cell.number_format = self.price_format
            
            ws.cell(row=current_row, column=10, value="").number_format = self.price_format
            ws.cell(row=current_row, column=11, value=f'=BDP(A{current_row},"PX_LAST")').number_format = self.price_format
            
            # Sensitivity price formulas
            ws.cell(row=current_row, column=14, 
                value=f"=IF($M$1<>\"\",I{current_row}*(1-$M$1/100),\"\")").number_format = self.price_format
            ws.cell(row=current_row, column=16,
                value=f"=IF($M$1<>\"\",I{current_row}*(1+$M$1/100),\"\")").number_format = self.price_format
            ws.cell(row=current_row, column=18,
                value=f"=IF($M$1<>\"\",K{current_row}*(1-$M$1/100),\"\")").number_format = self.price_format
            ws.cell(row=current_row, column=20,
                value=f"=IF($M$1<>\"\",K{current_row}*(1+$M$1/100),\"\")").number_format = self.price_format
            
            detail_rows = []
            current_row += 1

            # Sort positions within each underlying
            underlying_positions.sort(key=lambda x: (x.expiry_date, x.strike_price))

            # Write detail rows
            for pos in underlying_positions:
                ws.cell(row=current_row, column=2, value=pos.bloomberg_ticker)
                ws.cell(row=current_row, column=3, value=pos.expiry_date.strftime('%d/%m/%Y'))
                ws.cell(row=current_row, column=4, value=pos.position_lots)
                ws.cell(row=current_row, column=5, value=pos.security_type)
                
                strike_cell = ws.cell(row=current_row, column=6,
                    value=pos.strike_price if pos.strike_price > 0 else "")
                if pos.strike_price > 0:
                    strike_cell.number_format = self.price_format
                
                # Deliverable formulas
                for col_idx, price_col in [(7, "I"), (8, "J"), (12, "K")]:
                    formula = self._create_deliverable_formula(
                        current_row, group_start_row, pos.security_type, 
                        pos.strike_price, pos.position_lots, price_col
                    )
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                    cell.number_format = self.deliv_format
                
                # Sensitivity deliverable formulas
                for col_idx, price_col in [(15, "N"), (17, "P"), (19, "R"), (21, "T")]:
                    base_formula = self._create_deliverable_formula(
                        current_row, group_start_row, pos.security_type,
                        pos.strike_price, pos.position_lots, price_col
                    )
                    cell = ws.cell(row=current_row, column=col_idx, 
                        value=f"=IF($M$1<>\"\",{base_formula[1:]},\"\")")
                    cell.number_format = self.deliv_format
                
                for col in range(1, 22):
                    ws.cell(row=current_row, column=col).border = self.border
                
                detail_rows.append(current_row)
                current_row += 1
            
            # Group header totals
            if detail_rows:
                for col_idx in [7, 8, 12]:
                    col_letter = chr(64 + col_idx)
                    cell = ws.cell(row=group_start_row, column=col_idx,
                        value=f"=SUM({col_letter}{detail_rows[0]}:{col_letter}{detail_rows[-1]})")
                    cell.number_format = self.deliv_format
                
                for col_idx, col_letter in [(15, "O"), (17, "Q"), (19, "S"), (21, "U")]:
                    cell = ws.cell(row=group_start_row, column=col_idx,
                        value=f"=IF($M$1<>\"\",SUM({col_letter}{detail_rows[0]}:{col_letter}{detail_rows[-1]}),\"\")")
                    cell.number_format = self.deliv_format
            
            for row_num in detail_rows:
                ws.row_dimensions[row_num].outline_level = 1
                ws.row_dimensions[row_num].hidden = False
        
        self._set_master_column_widths(ws)
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.summaryRight = False
    
    def write_iv_master_sheet(self, positions: List[Position], prices: Dict[str, float]):
        """Write master IV sheet with all expiries"""
        ws = self.wb.create_sheet("IV_All_Expiries")
        
        # Grand totals row
        grand_total_row = 1
        ws.cell(row=grand_total_row, column=1, value="GRAND TOTAL").font = Font(bold=True, size=12)
        
        for col in range(1, 17):
            cell = ws.cell(row=grand_total_row, column=col)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.border = self.border
        
        # Headers
        headers = [
            "Underlying", "Symbol", "Expiry", "Position", "Type", "Strike", "Lot Size",
            "System IV (INR)", "System IV (USD)", "Override IV (INR)", "Override IV (USD)",
            "System Price", "Override Price", "BBG Price", "BBG IV (INR)", "BBG IV (USD)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Write grouped data
        grouped = self._group_positions(positions)
        sorted_underlyings = sorted(grouped.keys())
        
        current_row = 3
        
        for underlying in sorted_underlyings:
            underlying_positions = grouped[underlying]
            group_start_row = current_row
            
            # Group header
            cell = ws.cell(row=current_row, column=1, value=underlying)
            cell.font = self.group_font
            
            for col in range(1, 17):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = self.group_fill
                cell.border = self.border
            
            spot_price = prices.get(underlying)
            
            # Prices on group header
            price_cell = ws.cell(row=current_row, column=12, value=spot_price if spot_price else "")
            if spot_price:
                price_cell.number_format = self.price_format
            
            ws.cell(row=current_row, column=13, value="").number_format = self.price_format
            ws.cell(row=current_row, column=14, value=f'=BDP(A{current_row},"PX_LAST")').number_format = self.price_format
            
            detail_rows = []
            current_row += 1

            underlying_positions.sort(key=lambda x: (x.expiry_date, x.strike_price))

            # Write detail rows
            for pos in underlying_positions:
                ws.cell(row=current_row, column=2, value=pos.bloomberg_ticker)
                ws.cell(row=current_row, column=3, value=pos.expiry_date.strftime('%d/%m/%Y'))
                ws.cell(row=current_row, column=4, value=pos.position_lots)
                ws.cell(row=current_row, column=5, value=pos.security_type)

                strike_cell = ws.cell(row=current_row, column=6,
                    value=pos.strike_price if pos.strike_price > 0 else "")
                if pos.strike_price > 0:
                    strike_cell.number_format = self.price_format

                ws.cell(row=current_row, column=7, value=pos.lot_size)

                # IV formulas
                for col_idx, price_col in [(8, "L"), (10, "M"), (15, "N")]:
                    formula = self._create_iv_formula(current_row, group_start_row, price_col)
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                    cell.number_format = self.iv_format

                # USD conversions
                for inr_col, usd_col in [(8, 9), (10, 11), (15, 16)]:
                    inr_cell = chr(64 + inr_col)
                    cell = ws.cell(row=current_row, column=usd_col, 
                        value=f"={inr_cell}{current_row}/{self.usdinr_rate}")
                    cell.number_format = self.iv_format
                
                for col in range(1, 17):
                    ws.cell(row=current_row, column=col).border = self.border
                
                detail_rows.append(current_row)
                current_row += 1
            
            for row_num in detail_rows:
                ws.row_dimensions[row_num].outline_level = 1
                ws.row_dimensions[row_num].hidden = False
        
        # Grand totals
        for col_idx, col_letter in [(8, "H"), (9, "I"), (10, "J"), (11, "K"), (15, "O"), (16, "P")]:
            cell = ws.cell(row=grand_total_row, column=col_idx, 
                value=f"=SUM({col_letter}3:{col_letter}1000)")
            cell.number_format = self.iv_format
        
        self._set_iv_column_widths(ws)
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.summaryRight = False
    
    def write_iv_expiry_sheet(self, expiry_date: datetime, positions: List[Position], prices: Dict[str, float]):
        """Write IV sheet for specific expiry"""
        sheet_name = f"IV_Expiry_{expiry_date.strftime('%d_%m_%Y')}"
        ws = self.wb.create_sheet(sheet_name)
        
        expiry_positions = [p for p in positions if p.expiry_date.date() == expiry_date.date()]
        
        if not expiry_positions:
            return
        
        # Row 1: Grand totals row
        grand_total_row = 1
        ws.cell(row=grand_total_row, column=1, value="GRAND TOTAL").font = Font(bold=True, size=12)
        
        for col in range(1, 17):
            cell = ws.cell(row=grand_total_row, column=col)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.border = self.border
        
        # Row 2: Headers
        headers = [
            "Underlying", "Symbol", "Expiry", "Position", "Type", "Strike", "Lot Size",
            "System IV (INR)", "System IV (USD)", "Override IV (INR)", "Override IV (USD)",
            "System Price", "Override Price", "BBG Price", "BBG IV (INR)", "BBG IV (USD)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Group positions by underlying
        grouped = self._group_positions(expiry_positions)
        sorted_underlyings = sorted(grouped.keys())
        
        current_row = 3
        
        for underlying in sorted_underlyings:
            underlying_positions = grouped[underlying]
            group_start_row = current_row
            
            # Write group header
            cell = ws.cell(row=current_row, column=1, value=underlying)
            cell.font = self.group_font
            
            for col in range(1, 17):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = self.group_fill
                cell.border = self.border
            
            spot_price = prices.get(underlying)
            
            # Prices on group header
            price_cell = ws.cell(row=current_row, column=12, value=spot_price if spot_price else "")
            if spot_price:
                price_cell.number_format = self.price_format
            
            ws.cell(row=current_row, column=13, value="").number_format = self.price_format
            ws.cell(row=current_row, column=14, value=f'=BDP(A{current_row},"PX_LAST")').number_format = self.price_format
            
            detail_rows = []
            current_row += 1
            
            # Sort positions within each underlying
            underlying_positions.sort(key=lambda x: (x.expiry_date, x.strike_price))

            # Write detail rows
            for pos in underlying_positions:
                ws.cell(row=current_row, column=2, value=pos.bloomberg_ticker)
                ws.cell(row=current_row, column=3, value=pos.expiry_date.strftime('%d/%m/%Y'))
                ws.cell(row=current_row, column=4, value=pos.position_lots)
                ws.cell(row=current_row, column=5, value=pos.security_type)

                strike_cell = ws.cell(row=current_row, column=6,
                    value=pos.strike_price if pos.strike_price > 0 else "")
                if pos.strike_price > 0:
                    strike_cell.number_format = self.price_format

                ws.cell(row=current_row, column=7, value=pos.lot_size)

                # IV formulas
                for col_idx, price_col in [(8, "L"), (10, "M"), (15, "N")]:
                    formula = self._create_iv_formula(current_row, group_start_row, price_col)
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                    cell.number_format = self.iv_format
                
                # USD conversions
                for inr_col, usd_col in [(8, 9), (10, 11), (15, 16)]:
                    inr_cell = chr(64 + inr_col)
                    cell = ws.cell(row=current_row, column=usd_col, 
                        value=f"={inr_cell}{current_row}/{self.usdinr_rate}")
                    cell.number_format = self.iv_format
                
                for col in range(1, 17):
                    ws.cell(row=current_row, column=col).border = self.border
                
                detail_rows.append(current_row)
                current_row += 1
            
            for row_num in detail_rows:
                ws.row_dimensions[row_num].outline_level = 1
                ws.row_dimensions[row_num].hidden = False
        
        # Grand totals
        for col_idx, col_letter in [(8, "H"), (9, "I"), (10, "J"), (11, "K"), (15, "O"), (16, "P")]:
            cell = ws.cell(row=grand_total_row, column=col_idx, 
                value=f"=SUM({col_letter}3:{col_letter}1000)")
            cell.number_format = self.iv_format
        
        self._set_iv_column_widths(ws)
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.summaryRight = False
    
    def write_all_positions_sheet(self, positions: List[Position]):
        """Write sheet with all positions as a simple dump"""
        ws = self.wb.create_sheet("All_Positions")
        
        headers = ["Underlying", "Symbol", "Expiry", "Position", "Type", "Strike", "Lot Size"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        sorted_positions = sorted(positions,
                                key=lambda x: (x.underlying_ticker, x.expiry_date, x.strike_price))

        current_row = 2
        for pos in sorted_positions:
            ws.cell(row=current_row, column=1, value=pos.underlying_ticker)
            ws.cell(row=current_row, column=2, value=pos.bloomberg_ticker)
            ws.cell(row=current_row, column=3, value=pos.expiry_date.strftime('%d/%m/%Y'))
            ws.cell(row=current_row, column=4, value=pos.position_lots)
            ws.cell(row=current_row, column=5, value=pos.security_type)
            
            strike_cell = ws.cell(row=current_row, column=6, 
                                value=pos.strike_price if pos.strike_price > 0 else "")
            if pos.strike_price > 0:
                strike_cell.number_format = self.price_format
            
            ws.cell(row=current_row, column=7, value=pos.lot_size)
            
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).border = self.border
            
            current_row += 1
        
        widths = {
            'A': 25, 'B': 30, 'C': 12, 'D': 10, 'E': 8, 'F': 10, 'G': 10
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    
    def write_unmapped_sheet(self, unmapped_symbols: List[Dict]):
        """Write sheet with unmapped symbols"""
        ws = self.wb.create_sheet("Unmapped_Symbols")
        
        headers = ["Symbol", "Expiry", "Position Lots"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        
        for row, item in enumerate(unmapped_symbols, 2):
            ws.cell(row=row, column=1, value=item['symbol'])
            ws.cell(row=row, column=2, value=item['expiry'].strftime('%d/%m/%Y'))
            ws.cell(row=row, column=3, value=item['position_lots'])
    
    def save(self):
        """Save the Excel file"""
        self.wb.save(self.output_file)
        logger.info(f"Saved Excel file: {self.output_file}")
    
    # Helper methods
    def _group_positions(self, positions: List[Position]) -> Dict[str, List[Position]]:
        """Group positions by underlying"""
        grouped = {}
        for pos in positions:
            if pos.underlying_ticker not in grouped:
                grouped[pos.underlying_ticker] = []
            grouped[pos.underlying_ticker].append(pos)
        
        for underlying in grouped:
            grouped[underlying].sort(key=lambda x: (x.expiry_date, x.strike_price))
        
        return grouped
    
    def _create_deliverable_formula(self, row: int, group_header_row: int, 
                                   security_type: str, strike: float, lots: float, 
                                   price_column: str) -> str:
        """Create Excel formula for deliverable calculation"""
        type_cell = f"E{row}"
        position_cell = f"D{row}"
        strike_cell = f"F{row}"
        price_cell = f"${price_column}${group_header_row}"
        
        formula = (
            f'=IF({type_cell}="Futures",{position_cell},'
            f'IF({type_cell}="Call",IF({price_cell}>{strike_cell},{position_cell},0),'
            f'IF({type_cell}="Put",IF({price_cell}<{strike_cell},-{position_cell},0),0)))'
        )
        
        return formula
    
    def _create_iv_formula(self, row: int, group_header_row: int, price_column: str) -> str:
        """Create Excel formula for IV calculation"""
        type_cell = f"E{row}"
        position_cell = f"D{row}"
        strike_cell = f"F{row}"
        lot_size_cell = f"G{row}"
        price_cell = f"${price_column}${group_header_row}"
        
        formula = (
            f'=IF({type_cell}="Futures",0,'
            f'IF({type_cell}="Call",IF({price_cell}>{strike_cell},{position_cell}*{lot_size_cell}*({price_cell}-{strike_cell}),0),'
            f'IF({type_cell}="Put",IF({price_cell}<{strike_cell},{position_cell}*{lot_size_cell}*({strike_cell}-{price_cell}),0),0)))'
        )
        
        return formula
    
    def _write_deliverable_sheet_content(self, ws: Worksheet, positions: List[Position], prices: Dict[str, float]):
        """Helper to write deliverable sheet content - used by expiry sheets"""
        # Implementation follows same pattern as write_master_sheet
        # This is a helper to avoid code duplication
        pass
    
    def _write_iv_sheet_content(self, ws: Worksheet, positions: List[Position], prices: Dict[str, float]):
        """Helper to write IV sheet content - used by IV expiry sheets"""
        # Implementation follows same pattern as write_iv_master_sheet
        # This is a helper to avoid code duplication
        pass
    
    def _set_master_column_widths(self, ws: Worksheet):
        """Set column widths for master sheet"""
        widths = {
            'A': 25, 'B': 30, 'C': 12, 'D': 10, 'E': 8, 'F': 10,
            'G': 15, 'H': 15, 'I': 12, 'J': 12, 'K': 12, 'L': 15,
            'M': 8, 'N': 14, 'O': 14, 'P': 14, 'Q': 14, 'R': 14,
            'S': 14, 'T': 14, 'U': 14
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    
    def _set_iv_column_widths(self, ws: Worksheet):
        """Set column widths for IV sheets"""
        widths = {
            'A': 25, 'B': 30, 'C': 12, 'D': 10, 'E': 8, 'F': 10, 'G': 10,
            'H': 15, 'I': 15, 'J': 15, 'K': 15, 'L': 12, 'M': 12, 'N': 12,
            'O': 15, 'P': 15
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width